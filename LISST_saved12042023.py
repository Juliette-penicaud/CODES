#31/03/2023 Objectif : représenter les données du LISST : SPMVS, PSD, D50
#Idée : relier val surface au % MO ==> Con-in firmer agglomération de particule via TEP

import pandas as pd
import numpy as np
import sys, os
import matplotlib as mpl
import matplotlib.pyplot as plt
from datetime import datetime
import scipy.signal as signal

bidouille1=0
bidouille=True
if bidouille1 : #aller jusqua la derniere valeur de la valeur que l'on choisi
    rep='/home/penicaud/Documents/Modèles/Création_bathy/'
    file=rep+'post_smoothed_bathy_finale_v5.ijDeltaH'
    outfile=rep+'post_smoothed_bathy_finale_v5_unique.ijDeltaH'
    data=pd.read_csv(file, sep=' ', header=None)
    print(data)
    #df_sortie = pd.DataFrame(data)
    data.columns = ['i', 'j', 'depth', 'h']
    df_sortie_unique = data.drop_duplicates(subset=['i', 'j'], keep='last')
    df_sortie_unique.to_csv(outfile, sep=' ', index=False, float_format='%.4f')
    print(outfile)
    print(df_sortie_unique)
    sys.exit(1)
elif bidouille :
    rep='/home/penicaud/Documents/Modèles/Création_bathy/'
    file=rep+'partie_bathy3ple_+2m_v2.ijDeltaH'
    outfile=rep+'partie_bathy3ple_+2m_modified.ijDeltaH'
    data=pd.read_csv(file, sep=' ', header=None)
    print(data)
    data.columns = ['x', 'y', 'depth', 'h']
    val_test=(data[data['y'] < 1.43 * data['x'] + 291.3])
    data = data.drop(np.where(data[data['y'] < 1.43 * data['x'] + 291.3])[0])
    val_test = val_test.drop(['depth', 'h'], axis=1)
    fig, ax = plt.subplots()
    ax.scatter(data['x'], data['y'], color='red', marker='x', alpha=0.5)
    ax.scatter(val_test['x'], val_test['y'], color='blue')
    fig.savefig('test_points_bathy')

    data.to_csv(outfile)

    sys.exit(1)

year='2022'
list_month= ['June','August','Octobre']
i = 0 # 0 1 2 a voir pour faire une boucle si besoin
month = list_month[i]
rep = '/home/penicaud/Documents/Data/LISST/'
suffixe='.xlsx'
file = rep+'LISST_'+month+'_'+year+suffixe
print('file : ', file)
create=False
save=True
test=True
SspeedD50=False

cmap=plt.cm.jet
d1=pd.ExcelFile(file)
list_sheet=d1.sheet_names
print(type(list_sheet))
#print((list_sheet[i][0]=='S' for i in range(0,len(list_sheet)-1)))
if month=='June':
    list2=pd.Series(list_sheet).str.contains('S') # <----
elif month == 'August':
    list2=pd.Series(list_sheet).str.contains('A')
#print(list2)
#print(list_sheet[np.where(list2==True)[0]])
indice=np.where(list2==True)[0]
print('indice', indice) #donne l'indice du sheet name à aller voir. On va boucler dessus

#Je charge une seule fois les valeurs des size class, comme ce sont les mêmes, n'importe quelle station est ok
#size_class = pd.read_excel(file, usecols='AV:BY', nrows=2)
#print(size_class)
#width_size_class = size_class.iloc[0]
#list_noms = list(size_class.columns)
#print(list_noms)

#Record the lowest depth
list_depth=[]

list_class=['1.25-1.48', '1.48-1.74', '1.74-2.05', '2.05-2.42', '2.42-2.86', '2.86-3.38', '3.38-3.98', '3.98-4.70',
            '4.70-5.55', '5.55-6.55', '6.55-7.72', '7.72-9.12', '9.12-10.8', '10.8-12.7', '12.7-15.0', '15.0-17.7',
            '17.7-20.9', '20.9-24.6', '24.6-29.1', '29.1-34.3', '34.3-40.5', '40.5-47.7', '47.7-56.3', '56.3-66.5',
            '66.5-78.4', '78.4-92.6', '92.6-109', '109-129', '129-152', '152-180', '180-212', '212-250']

#print(list_class[0][0:4], list_class[0][5:9])
#print(list_class[-5][0:3], list_class[-5][4:8])
list_width = []
list_median_size=[]
list_noms_col=[]
for i in range(len(list_class)):
    #print('i', i)
    list_noms_col.append('#'+str(i+1))
    if i < 27 :
        step=4
    else:
        step=3
    val1=float(list_class[i][step+1:2*step+1])
    val2=float(list_class[i][0:step] )
    val=val1-val2
    #print('val1',val1, 'val2', val2, '=', val)
    list_width.append(val)
    list_median_size.append(val2+val/2)

list_noms_col=list_noms_col+['laser trans', 'battery V', 'ext aux input', 'laser ref', 'P', 'T', 'int date', 'int date 2', 'inc 1', 'inc 2', 'Date', 'Hour']
print(list_noms_col)
print(len(list_noms_col))
print(list_width)
print(list_median_size)
list_median_size_name=[str(round(m,2)) for m in list_median_size]
print(len(list_width))
print(list_sheet[0])
print(indice)

fontsize=10
var = 'D50'  # 'depth or D50 # TO CHANGE
if var == 'depth':
    un_var = ' (m)'
    vmin=-15
    vmax=0
elif var == 'D50':
    un_var = ' (µm)'
    vmin=0
    vmax=170
else :
    print('ERROR choose good var for the Diagram plot')
    sys.exit(1)

if SspeedD50 :
    figSspeedD50, axSspeedD50 = plt.subplots(ncols=1)
    figSspeedD50.suptitle('Evolution of '+var+' depending on velocity and salinity. Survey ' + month, fontsize=fontsize)
    axSspeedD50.set_xlabel('Salinity (PSU)', fontsize=fontsize)
    axSspeedD50.set_ylabel('Current velocity (mm/s)', fontsize=fontsize)


for i in indice:# len(indice)] : #indice #45:46
    station=str(list_sheet[i])
    print('station', station)
    #data_month=pd.read_excel(file, usecols='AK,AQ:DI,DN:EP', sheet_name='S2')#, header=0) #'AK,AQ:BY' BY to get all the values, but the extreme classes have been taken for the calculations
    if create :
        data_month=pd.read_excel(file, usecols='A:AR', sheet_name=station, skiprows=3)#, skiprows=3)#, header=0) #
        data_month=data_month.dropna() #how='any'
        data_month=data_month[(data_month != 0).all(axis=1)] #On enlève tous les 0
        #print(data_month.head(20))
        #print(data_month.columns)
        data_month.columns = list_noms_col
        #print(data_month.iloc[0]) #print la ligne 0
        #print(data_month.iloc[:]['#1']) # print la colonne de #1
        # 1e CRITERE de selection des station qui fonctionnent ou non
        if np.shape(data_month['P'].drop_duplicates().index.values)[0] <10 : #Skip si le fichier ne contient pas plus de 10 profondeurs différentes
            continue
        #Done  : critere de selectio, : si longueur fichier trop petite, skipper


        ##########################    DETECTER LE PROFIL DESCENDANT   ####################################""""
        #version 1 : je vais juste trouver la cellule la plus profonde et utiliser tout ce qu'il y a au dessus jusqu'au min
        min1=data_month['P'].min()
        list_depth.append(min1) #je choisis de l'ajouter dans tous les cas pour avoir toutes les données min, quitte à avoir plus de val min que de stations
        idxmin=data_month['P'].idxmin()
        max=data_month['P'].max()
        idxmax=data_month['P'].drop_duplicates(keep='last').idxmax()#data_month['P'].idxmax() # permet de trouver index de la dernière rep de cette valeur max
        if idxmin>idxmax: #Si val min détectée à la surface
            print('VAL MIN RECALCULEE')
            idxmin=data_month['P'].loc[0:idxmax].idxmin()
            min2=data_month['P'].loc[0:idxmax].min()
            list_depth.append(min2)
            min1=min(min1,min2)

        print(min1, 'index_min', idxmin, '\n',  max, idxmax )
        data_month2=data_month.loc[idxmin:idxmax].copy()
        #print(data_month2)
        # CRITERE 2 de selection des station qui fonctionnent ou non
        # if np.shape(data_month)[0]<15 : #si le fichier ou on a enlevé les 0 est trop court
        #     continue
        if np.shape(data_month2['P'].drop_duplicates().index.values)[0] <10 : #Skip si le fichier ne contient pas plus de 10 profondeurs différentes
            continue
        #Done  : critere de selectio, : si longueur fichier trop petite, skipper
        print('shape data 1 : ', np.shape(data_month))
        print('shape data 2 : ', np.shape(data_month2), 'diff', np.shape(data_month)[0]-np.shape(data_month2)[0] )
        #Je corrige P,   #Calculer les offsets  #


        if (min1 > 0.5 and month!='June'):  # TODO A VOIR SI ON FAIT CA QUE POUR AOUT OU POUR TOUTES LES CAMPAGNES
            print('aha')
            data_month2['Pcorr']=data_month2['P'].values-(min1-0.5)# on va considérer que première mesure est à 50cm sous surface
        else :
            print('Ohooh')
            data_month2['Pcorr']=data_month2['P']


        ########################################################################################################
        #Traiter les données
        classe_first=2
        classe_last=30
        df_process=data_month2.iloc[:,classe_first:classe_last] # Je choisis les colonnes qui correspondent aux classes 3 à 30 inclu
        prof=data_month2['P']
        #print('prof',prof)
        #df_process=pd.concat([df_process,prof])
        print('df_process',df_process)

        df_SPMVC=pd.DataFrame()
        df_D50=pd.DataFrame()
        df_SPMVC['Vol Tot']=df_process.sum(axis=1)
        df_SPMVC['P']=data_month2['P'].values
        #print(df_SPMVC, type(df_SPMVC))
        #print(list_width)

        for c in range(classe_first+1, classe_last+1): #boucle qui dépend de la taille de df_process, ie de 32 - classes extérieures virées, +1 car on utilise #1 et non indice 0
            classe="#"+str(c)
            #print(df_process.iloc[:][classe])
            df_SPMVC["SPMVC "+classe]=df_process.iloc[:][classe]/df_SPMVC.iloc[:]['Vol Tot'] #je calcule le SPMVC pour chaque classe
            if c-classe_first == 1 :
                df_D50[classe] = df_SPMVC["SPMVC "+classe]*100 #SPMVC cumulé : égal à SMPVC #1
            else :
                df_D50[classe] = df_D50["#" + str(c-1)] +(df_SPMVC["SPMVC "+classe])*100 #SPMVC cumulé : égal à SMPVC de la classe + le cumulé de la classe précédante
            df_SPMVC["PSD "+classe]=df_SPMVC["SPMVC "+classe]/list_width[c-1]  #c-1 car redevient un indice pour la list_width
            # #SPMVC normalisé par la taille de la classe des particules
            #calcul du D50 à partir du
            #print(df_SPMVC[classe])


        #Calcul du D50 : trouver ou est le plus proche de 50 tout en étant inféreur
        print(df_D50)

        df_test = (df_D50.T) <= 50 #je sélectionne les valeurs inf ou égale à 50, renvoie tab boolean
        list_D50=df_test[:].idxmin() #je détecte la première valeur a false, donne l'index
        list_classe_D50, list_classe_D51=[], []
        list_val_classe_D50, list_val_classe_D51=[], []
        c=0
        for i in range(len(list_D50)):
            #end=len(list_D50[i])
            ind_classe=int(list_D50.iloc[i][1:3])-2 #J'extrai sous la forme d'int l'INDEX de la classe que l'on a trouvé avec list_D50 (numéro =num+1)
            list_classe_D50.append(ind_classe) #list des valeurs des classes, à faire -1 pour aller chercher la valeur de la taille de la classe pour le calcul suivant
            #-1 car l'indice donne la première valeur à false, et non la dernière à true
            list_classe_D51.append(ind_classe+1)
            val_classe=df_D50.iloc[c].iloc[ind_classe - classe_first] #je selectionne ligne par ligne (c) la position détectée
            # de la colonne qui correspond à <50, numéro-val de la 1ere col qu'on a retiré, -1 car on est en index
            val_classe1 = df_D50.iloc[c].iloc[ind_classe+1 - classe_first]
            list_val_classe_D50.append(val_classe)
            list_val_classe_D51.append(val_classe1)
            c=c+1
        print('list D50 val classe', list_classe_D50)
        print('val classe', list_val_classe_D50)

        df_mediansize = pd.Series(list_median_size)

        reste_D50 = pd.DataFrame()#list_val_classe_D50)
        reste_D50['val 50'] = list_val_classe_D50
        reste_D50['reste']= 50 - reste_D50['val 50']
        reste_D50['val 51'] = list_val_classe_D51
        reste_D50['diff val'] = reste_D50['val 51'] - reste_D50['val 50']
        reste_D50['median class 50'] = df_mediansize[list_classe_D50].values
        reste_D50['median class 51'] = df_mediansize[list_classe_D51].values
        reste_D50['diff median']=reste_D50['median class 51']-reste_D50['median class 50']
        reste_D50['value finale']=reste_D50['median class 50']+reste_D50['reste']*(reste_D50['diff median']/reste_D50['diff val'])
        reste_D50['P'] = data_month2['P'].values
        print(reste_D50)

        df_SPMVC['D50']=reste_D50['value finale'].values
        df_SPMVC['Date']=data_month2['Date'].copy()
        df_SPMVC['Hour'] = data_month2['Hour'].copy()
        df_SPMVC['Pcorr'] = data_month2['Pcorr'].copy()

        if save:
            outfile=rep+'Survey_'+month+'/'+station+'_SPM'
            #reste_D50.to_csv(outfile+'_D50', sep=' ', index=False, float_format='%.4f')
            df_SPMVC.to_csv(outfile, sep=' ', index=False, float_format='%.4f')
            print(outfile)
        # faire des moyennes ?
        # COMMENT choisir les couches ? Tout par mètres ? Sachant que problème d'offset.
        # Sur un mètre ? Checker la disparités via STD ?
        #df_process['P'] = prof  # J'ajoute une colonne au dataframe pour pouvoir faire des moyennes
        #moy_2m = df_process[df_process['P'] < 2]
        #print(moy_2m)

    else :
        #CHARGER fichier
        file_data=rep+'Survey_'+month+'/'+station+'_SPM'
        print('open file', file_data)
        if not os.path.exists(file_data):
            print('I skip '+file_data+' it does not exist')
            continue
        data=pd.read_csv(file_data, sep=' ')
        pmin=data['Pcorr'].min()
        pmax=data['Pcorr'].max()
        print('pmax', pmax)

        #MOYENNE par valeurs de P identiques
        list_prof = list(set(data['Pcorr'].values))
        list_prof.sort()
        list_average_depth=[]
        data_unique=pd.DataFrame()
        c=0

        hour=data['Hour'].loc[1]
        date=data['Date'].loc[1]
        date=datetime.strptime(date+' '+hour, '%Y-%m-%d %H:%M:%S')
        #to avoid error i drop the Date and Hour columns
        data=data.drop(["Date", 'Hour'], axis=1)
        print('date', date, hour)

        for l in list_prof :
            print('prof étudiée = ',l)
            #print(data['D50'].loc[data['Pcorr']==l])
            #print('STD sur D50', data['D50'].loc[data['Pcorr']==l].std(ddof=0))
            new_row=pd.DataFrame(data.loc[data['Pcorr'] == l].median())
            #print('new row ok')
            data_unique=pd.concat([data_unique,new_row.T], ignore_index=True)
            #print( 'concat ok ')
            data_unique.loc[c, 'STD D50']=data['D50'].loc[data['Pcorr']==l].std(ddof=0) #donne la std de toutes les valeurs à la profondeur etudiée
            #print('add of STD val ok')
            c=c+1
        print('ok for data_unique')

        #MOYENNE par couche #TODO : est ce qu'il vaut mieux faire ca a partir de data_unique (median) ou une moyenne de toutes les couches ?
        data_averaged=pd.DataFrame() #QUESTION : tous les m ? 2m ? Divisé en 3 (sur,bot,mid ?)
        #Je teste en faisant un moyenne tous les metres
        p=0
        if pmax <= 3 :
            print('ATTENTION VAL MAX < 3m')#, att aux moy de surf et bott')
        for i in range(int(pmax)+1):
            new_row = pd.DataFrame((data.loc[(data['Pcorr'] > p) & (data['Pcorr'] < p+1) ]).mean()) # TODO il faudra mettre P corr
            #TODO : quoi faire ? Avec median i.e data unique ou mean i.e avec data ?
            data_averaged = pd.concat([data_averaged, new_row.T], ignore_index=True)
            data_averaged.loc[i,'nb val']= np.shape(data[(data['Pcorr'] < 2) == True])[0]
            data_averaged.loc[i, 'STD D50'] = data['D50'].loc[data['Pcorr'] <= p+1].std(ddof=0) # donne la std de la moyenne de prof avec toutes les valeurs inclues
            p=p+1
        #moy_first_m = data[(data['P'] < 2) == True]
        #list_average_depth.append(moy_first_m)
        #moy_last_m =  data[(data['P'] > pmax-2) == True] #average over 2m over bot
        #list_average_depth.append(moy_last_m)
        #print('len moy first and last', np.shape(moy_first_m)[0],np.shape(moy_last_m)[0] )
        #if pmax>5 :
        #    moy_mid_m = data[((data['P'] >= 2)& (data['P'] <= pmax-2) )== True]
        #    list_average_depth.append(moy_mid_m) #TODO si supérieur à x m, on fait plusieurs couches


        fontsize=10
        outfile = rep + 'Survey_' + month + '/figure/'
        if not os.path.exists(outfile+'/'+station):
            # if the demo_folder directory is not present
            # then create it.
            os.makedirs(outfile+'/'+station)
        outfile=outfile+'/'+station+'/'

        fixed = (station[0:3] == 'S40') or (station[0:2] == 'AF') # or for octobre #CHECK IF FIXED sta or not

        if not test :
            ###################    figure de l'évolution de D50 % depth
            fig, ax = plt.subplots(ncols=1)
            if fixed :
                fig.suptitle('D50 Fixed station '+date.date()+' '+date.time()+' '+hour)
            else :
                fig.suptitle('D50 Survey '+month+', '+station)
            ax.scatter(data_unique['D50'], -data_unique['Pcorr'], marker='x', alpha=0.8, color='grey')
            ax.set_xlabel('D50', fontsize=fontsize)
            ax.set_ylabel('Depth (m)', fontsize=fontsize)
            fig.savefig(outfile + station+'_D50', format='png')

            ################## figure histogramme des PSD à chaque prof.
            data_PSD=data_unique.iloc[:, 3:-1:2]
            for d in range(np.shape(data_PSD)[0]):
                fig2,ax=plt.subplots()
                depth=str(data_unique['Pcorr'].loc[d])
                fig2.suptitle('Survey '+month+', '+station+' at depth='+depth+' m')
                #ax = data_PSD.plot.bar() #va ploter toutes les PSD pour chaque prof.
                ax = data_PSD.loc[d].plot.bar() #va ploter toutes les PSD pour UNE prof.
                ax.set_xlabel('Median size (µm)', fontsize=fontsize)
                #ax.xaxis.set_minor_locator(MultipleLocator(5))
                ax.set_xticks(np.arange(1,len(list_median_size_name[2:30])+1,1))
                ax.set_xticklabels(list_median_size_name[2:30],fontsize=fontsize-2, rotation = 30) #TODO : a atuomatiser avec classe first classe last
                ax.set_ylabel('PSD ($µl^{-1}$ µm)', fontsize=fontsize)
                ax.set_ylim(0,(int(100*data_PSD.loc[:].max().max())+1)/100)
                fig2.savefig(outfile+station+'_rep_PSD_at'+depth, format='png')

            #########################  Figure averaged depth
            #nom_list_average_depth=['first', 'last', 'intermediate'] #on sait que rangé comme : first_m, last_m, middle_m1; middle metre 2 ...
            data_PSD_averaged=data_averaged.iloc[:,3:-2:2]
            for i in range(np.shape(data_averaged)[0]):#len(list_average_depth)):
                fig3, ax = plt.subplots()
                d = data_PSD_averaged.loc[i]
                #d = list_average_depth[l].iloc[:,3:-1:2]
                if station[0:3]=='S40':
                    title= 'Survey ' + month + ', fixed station ' + str(hour) + ' Average around ' + str(i+1) + ' meter'
                else :
                    title = 'Survey ' + month + ', ' + station + ' Average around ' + str(i+1) + ' meter'
                fig3.suptitle(title, fontsize=fontsize)
                # ax = data_PSD.plot.bar() #va ploter toutes les PSD pour chaque prof.
                ax = d.plot.bar()  # va ploter toutes les PSD pour UNE prof moyennée
                ax.set_xlabel('Median size (µm)', fontsize=fontsize)
                #ax.xaxis.set_minor_locator(MultipleLocator(5))
                ax.set_xticks(np.arange(0,len(list_median_size_name[2:30]),1))
                ax.set_xticklabels(list_median_size_name[2:30],fontsize=fontsize-2, rotation=30) #TODO : a atuomatiser avec classe first classe last
                ax.set_ylabel('PSD ($µl^{-1}$ µm)', fontsize=fontsize)
                ax.set_ylim(0,(int(100*data_PSD.loc[:].max().max())+1)/100)
                fig3.savefig(outfile+station+'_PSD_averaged_'+str(i+1)+'m', format='png')

                # ###################    figure de l'évolution de D50 % depth
                # Not done cause the not averaged is the best, we do not lose information
                # fig, ax = plt.subplots(ncols=1)
                # fig.suptitle('D50 Survey '+month+', '+station)
                # ax.scatter(data_averaged['D50'], -data_averaged['Pcorr'], marker='x', alpha=0.8, color='grey')
                # ax.set_xlabel('D50', fontsize=fontsize)
                # ax.set_ylabel('Depth (m)', fontsize=fontsize)
                # fig.savefig(outfile + station+'_D50_averaged', format='png')

        else :
            #TODO : figure du D50 avec prof ET profil de vitesse
            # 1. lire l'heure et le jour depuis le fichier.
            # 2. aller chercher l'equivalent et la moyenne des X bins dans le fchier ADCP pour le plotter.
            rep_adcp='/home/penicaud/Documents/Data/ADCP/Survey_'+month+'/'
            rep_CTD = '/home/penicaud/Documents/Data/CTD/Survey_'+month+'/'
            suffixe='.csv'
            if month=='June':
                ##############################   DONNEE ADCP #####################################################
                taille_bin=0.3
                offset=0.6
                if date.day==16:
                    file_adcp=rep_adcp+str(date.day)+'0'+str(date.month)+'_alldata_BT'
                elif date.day==17 :
                    if int(station[1:3])<=28:
                        file_adcp = rep_adcp + str(date.day)+'0' + str(date.month) + '_T3_alldata_BT'
                    else :
                        file_adcp = rep_adcp + str(date.day)+'0' + str(date.month) + '_T4_alldata_BT'
                elif date.day==18:
                    file_adcp = rep_adcp + str(date.day)+'0' + str(date.month) + '_alldata_BT'

                ############################### DONNEES CTD IMER ###################################
                file_imer = rep_CTD+'CTD_VU_16-20_06.xlsx'


            file_adcp=file_adcp+suffixe
            print('file_adcp', file_adcp)
            data_adcp= pd.read_csv(file_adcp, skiprows=11, low_memory=False, sep=' ')  # , usecols=col_list)
            #Je cherche la ou correspond heure des stations
            data_profil = pd.DataFrame((data_adcp.loc[(data_adcp['HH'] ==date.hour) & (data_adcp['MM']==date.minute)]).mean(numeric_only=True))
            data_profil=data_profil.T
            mag_cols = [col for col in data_profil.columns if 'Mag' in col] #selection de seulement les colonnes voulues de Magnitude
            data_profil2=data_profil[mag_cols].copy()
            #data_profil2=pd.DataFrame(data_profil2.dropna())
            depth_adcp=pd.DataFrame((float(data_profil2.columns[i][len(data_profil2.columns[i])-2:len(data_profil2.columns[i])])-1)*taille_bin+offset for i in range(np.shape(data_profil2)[1]))
            #ici j'ai construit la ligne des profondeurs à partir de taille_bin (dépend de chaque mois) et du numéro du bin concerné : 'Mag, mm/s, 3' = taille_bin*(3-1)+offset
            d_adcp=data_profil2.T
            d_adcp['depth'] = depth_adcp.values
            d_adcp=d_adcp.rename(columns={0: 'vitesse'})
            #data_profil3 = signal.medfilt(data_profil2.T)  # filter 3
            #TODO resolve it with spatial filtering averaged ??
            list_depth_adcp = list(round(n, 1) for n in d_adcp['depth'].values)


            ############################   DONNEE DE SALINITÉ AVEC LA CTD IMER "#############################
            col_list_imer = ["Depth", "Temp", "Salinity", "Density", "Chl", "Turbidity"]
            #TODO : skip si sheet=station n'existe pas
            from openpyxl import load_workbook
            wb = load_workbook(file_imer, read_only=True)  # open an Excel file and return a workbook

            if station in wb.sheetnames:
                print('sheet1 exists')
                data_imer = pd.read_excel(file_imer, station, skiprows=23, usecols=col_list_imer)  # lambda x : x > 0 and x <= 27 )#, usecols=col_list)
                print(data_imer)
            else:
                data_imer=pd.DataFrame(np.nan,index=range(len(list_depth_adcp)),columns=['Salinity', 'Depth'] )

            ############## CREATE a file with all values at same depth
            sal_extract=pd.DataFrame()
            D50_extract=pd.DataFrame()
            for l in range(len(list_depth_adcp)):
                dmax = list_depth_adcp[l]
                if l != 0:
                    dmin = list_depth_adcp[l - 1]
                else:
                    dmin = 0
                #print('dmin et max', dmin, dmax) #depth imer : every 0.1m
                new_row = pd.DataFrame((data_imer.loc[(data_imer['Depth'] <= dmax) & (data_imer['Depth'] >= dmin)]).mean())  # moyennage ?
                # new_row = pd.DataFrame(data_imer.loc[(data_imer['Depth'] == list_depth_adcp)]) #extraire la valeur sans aucun traitement : pas de moyennage
                new_row_D50 = pd.DataFrame((data_unique.loc[(data_unique['Pcorr'] <= dmax) & (data_unique['Pcorr'] >= dmin)]).mean() ) # moyennage ?
                D50_extract = pd.concat([D50_extract, new_row_D50.T], ignore_index=True)
                sal_extract = pd.concat([sal_extract, new_row.T], ignore_index=True)
                #print('dmoy', sal_extract['Depth'][l])

            d_adcp['Salinity'] = sal_extract['Salinity'].values #on ajoute les données de salinité
            d_adcp['D50']=D50_extract['D50'].values
            #Do the same for D50


            ###################    figure de l'évolution de D50 % depth AVEC LE PROFIL DE VITESSE
            fig, ax = plt.subplots(ncols=1)
            if fixed:
                fig.suptitle('D50 Fixed station ' + str(date.date()) + ' ' + str(date.time()) + ' ' + hour, fontsize=fontsize)
            else:
                fig.suptitle('D50 Survey ' + month + ', ' + station, fontsize=fontsize)
            ax.scatter(data_unique['D50'], -data_unique['Pcorr'],  marker='x', alpha=0.8,  color='grey') #c=data_imer['Salinity'] ,cmap=cmap)#
            ax2 = ax.twiny()
            ax2.plot(d_adcp['vitesse'],-d_adcp['depth'],alpha=0.8,  color='orange' ) #Ajouter la salinité en couleur ? cmap=data_imer['Salinity'] ne marche pas mais essayer
            ax2.set_ylabel('Current speed (mm/s)', fontsize=fontsize)
            ax.set_xlim(0,140)
            ax2.set_xlim(100,900)
            ax.set_ylim(-15,0)
            ax.set_xlabel('D50', fontsize=fontsize)
            #ax.set_xlim(0,np.max())
            ax.set_ylabel('Depth (m)', fontsize=fontsize)
            fig.savefig(outfile + station + '_D50_fonction_vitesse', format='png')
            #TODO : Ajouter le profil de salinité pour faire speed vs S et colorbar de D50

            ###################    figure de l'évolution de D50 AVEC LE PROFIL DE VITESSE ET SALINITE
            if SspeedD50 :
                cmap=plt.cm.jet
                if var=='depth':
                    c_plot=-d_adcp[var]
                else :
                    c_plot=d_adcp[var]
                im= axSspeedD50.scatter(d_adcp['Salinity'], d_adcp['vitesse'] , c=c_plot, alpha=0.8, cmap=cmap, vmin=vmin, vmax=vmax)

if SspeedD50:
    cbar = plt.colorbar(im, label=var+un_var, ax=axSspeedD50)#, ticks=1)
    figSspeedD50.savefig(rep+'Survey_'+month+'/figure/Diagramm_Speed_Sal_'+var, format='png')


print('end    ')