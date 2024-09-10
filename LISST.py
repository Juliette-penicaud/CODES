# 31/03/2023 Objectif : représenter les données du LISST : SPMVS, PSD, D50
# Idée : relier val surface au % MO ==> Con-in firmer agglomération de particule via TEP

import pandas as pd
import numpy as np
import sys, os
import matplotlib as mpl
import matplotlib.pyplot as plt
import datetime
import scipy.signal as signal
from openpyxl import load_workbook
from scipy.optimize import curve_fit

# 27/05 : je créé les tableaux station_SPM_#2-#31 SANS enlever les 0 afin d'avoir val max de prof et de faire
# une comp des profondeurs dans le tableaau Recap_all (Create_excel_fil_all_data.py)
# pour Aout et Oct avec CTD et ADCP, en supposant en 1e lieu que la correction Pcorr[min] = 0.5



year = '2022'
list_month = ['June', 'August', 'Octobre']
i = 1  # 0 1 2 a voir pour faire une boucle si besoin
month = list_month[i]
rep = '/home/penicaud/Documents/Data/LISST/'
suffixe = '.xlsx'
file = rep + 'LISST_' + month + '_' + year + suffixe
print('file : ', file)
create = False
save = False
test_fill_surfer_file = False
SspeedD50 = False
figure = False
test_depth = False



cmap = plt.cm.jet
d1 = pd.ExcelFile(file)
list_sheet = d1.sheet_names
print(type(list_sheet))
# print((list_sheet[i][0]=='S' for i in range(0,len(list_sheet)-1)))
if month == 'June':
    list2 = pd.Series(list_sheet).str.contains('S')
elif month == 'August':
    list2 = pd.Series(list_sheet).str.contains('A')
elif month == 'Octobre':
    list2 = pd.Series(list_sheet).str.contains('O')

# print(list2)
# print(list_sheet[np.where(list2==True)[0]])
indice = np.where(list2 == True)[0]
print('indice', indice)  # donne l'indice du sheet name à aller voir. On va boucler dessus

# Je charge une seule fois les valeurs des size class, comme ce sont les mêmes, n'importe quelle station est ok
# size_class = pd.read_excel(file, usecols='AV:BY', nrows=2)
# print(size_class)
# width_size_class = size_class.iloc[0]
# list_noms = list(size_class.columns)
# print(list_noms)

# Record the lowest depth
list_depth = []

list_class = ['1.25-1.48', '1.48-1.74', '1.74-2.05', '2.05-2.42', '2.42-2.86', '2.86-3.38', '3.38-3.98', '3.98-4.70',
              '4.70-5.55', '5.55-6.55', '6.55-7.72', '7.72-9.12', '9.12-10.8', '10.8-12.7', '12.7-15.0', '15.0-17.7',
              '17.7-20.9', '20.9-24.6', '24.6-29.1', '29.1-34.3', '34.3-40.5', '40.5-47.7', '47.7-56.3', '56.3-66.5',
              '66.5-78.4', '78.4-92.6', '92.6-109', '109-129', '129-152', '152-180', '180-212', '212-250']

# print(list_class[0][0:4], list_class[0][5:9])
# print(list_class[-5][0:3], list_class[-5][4:8])
list_width = []
list_median_size = []
list_noms_col = []
for i in range(len(list_class)):
    # print('i', i)
    list_noms_col.append('#' + str(i + 1))
    if i < 27: # to record the values, need to record the number of digits needed to read
        step = 4  # curseur de lecture des valeurs, si list_class[i][] ne contient pas 4 val : i.e > 100=3  char, 92.6=4
    else:
        step = 3
    val1 = float(list_class[i][step + 1:2 * step + 1])
    val2 = float(list_class[i][0:step])
    val = val1 - val2
    # print('val1',val1, 'val2', val2, '=', val)
    list_width.append(val)
    list_median_size.append(val2 + val / 2)

list_noms_col = list_noms_col + ['laser trans', 'battery V', 'ext aux input', 'laser ref', 'P', 'T', 'int date',
                                 'int date 2', 'inc 1', 'inc 2']  # , 'Date', 'Hour']
print(list_noms_col)
print(len(list_noms_col))
print(list_width)
print(list_median_size)
list_median_size_name = [str(round(m, 2)) for m in list_median_size]
print(len(list_width))
print(list_sheet[0])
print(indice)

# Je charge les données des stations !
rep_station = '/home/penicaud/Documents/Data/Survey_' + month + '/'
file_station = rep_station + 'Stations_' + month + '.xlsx'
print('file stations recap', file_station)
data_station = pd.read_excel(file_station, usecols=['Stations', 'Time', 'Distance'])

fontsize = 10

if SspeedD50:
    figSspeedD50, axSspeedD50 = plt.subplots(ncols=1)
    figSspeedD50.suptitle('Evolution of D50 with velocity and salinity. Survey ' + month, fontsize=fontsize)
    axSspeedD50.set_xlabel('Salinity (PSU)', fontsize=fontsize)
    axSspeedD50.set_ylabel('Current velocity (mm/s)', fontsize=fontsize)

if test_depth:
    figdepth, axdepth = plt.subplots(ncols=1)
    figdepth.suptitle('Evolution of D50 with velocity and salinity. Survey ' + month, fontsize=fontsize)
    figdepth.set_xlabel('Station', fontsize=fontsize)
    figdepth.set_ylabel('Depth (m)', fontsize=fontsize)

classe_first = 1  # commencera a val+1
classe_last = 31  # finira a val

def func(x, a, s):
    return a * np.power(x, -s)  # forme n(D)=KD^-s


def Junge_parameter(l, D):
    n = 7
    lbis = l[n::]
    Dbis = D.values[n::]
    popt, pcov = curve_fit(func, lbis, Dbis)
    return popt


def plot_Junge_fit(l, D):
    popt = Junge_parameter(l, D)
    fig, ax = plt.subplots(ncols=1)
    ax.grid(True, alpha=0.5)
    plt.plot(l, func(l, *popt), 'r-',
             label='%5.3f $x^{%5.3f}' % tuple(popt))
    plt.loglog(l, D)
    ax.set_ylabel('Number of particles $L^{-1}$ $µm^{-1}$', fontsize=fontsize)
    ax.set_xlabel('Particle size (µm)', fontsize=fontsize)
    plt.legend()
    fig.savefig('test_Junge')


def calcul_date(date_1, date_2, year):  # 26/05/2023 : derive datetime form the columns and not from the data
    day_from = (date_1 / 100).astype(int)
    start_date = datetime.datetime(int(year), 1, 1)
    # add the number of days to the start date
    result_list = []
    for d in range(len(day_from)):
        result_date = start_date + datetime.timedelta(days=int(day_from[d] - 1))
        result_list.append(result_date)
    day = result_list
    # day = day_to_date(day_from,int(year))
    hour = ((date_1 / 100 - (date_1 / 100).astype(
        int)) * 100)  # WARNING : +1 necessary to have right value if astype(int) at the end
    mn = (date_2 / 100)
    sec = (((date_2 / 100) - (date_2 / 100).astype(int)) * 100)  # .astype(int)
    date_list = []
    for d in range(len(day)):
        date = datetime.datetime(int(year), day[d].month, day[d].day, int(hour[d]), int(mn[d]), int(sec[d]))
        date_list.append(date)
    return date_list


for i in indice[0:10]:  # len(indice)] : #indice #45:46
    station = str(list_sheet[i])
    print('station', station)
    # data_month=pd.read_excel(file, usecols='AK,AQ:DI,DN:EP', sheet_name='S2')#, header=0) #'AK,AQ:BY' BY to get all the values, but the extreme classes have been taken for the calculations
    if create:
        data_month = pd.read_excel(file, usecols='A:AP', sheet_name=station, skiprows=3)  # , skiprows=3)#, header=0) #
        # TODO !!!!! PB avec AR, car c'est fait a la main, et ne va pas jusque en bas des données, il faut calculer la date ap des données.
        data_month = data_month.dropna()  # [0:data_month.last_valid_index()]
        # in order to avoid nan value everywhere, because nan values cannot be understood in next calculations
        data_month.columns = list_noms_col
        data_month = data_month.reset_index()

        date = calcul_date(data_month['int date'], data_month['int date 2'], int(year))
        date = date[0]

        data_month = data_month.dropna()  # how='any'
        # print(data_month.head(20))
        # print(data_month.columns)
        # data_month.columns = list_noms_col
        # print(data_month.iloc[0]) #print la ligne 0
        # print(data_month.iloc[:]['#1']) # print la colonne de #1
        # 1e CRITERE de selection des station qui fonctionnent ou non
        if np.shape(data_month['P'].drop_duplicates().index.values)[0] < 10:
            # Skip si le fichier ne contient pas plus de 10 profondeurs différentes
            continue
        # Done  : critere de selectio, : si longueur fichier trop petite, skipper

        ##########################    DETECTER LE PROFIL DESCENDANT   ####################################""""
        # version 1 : je vais juste trouver la cellule la plus profonde et utiliser tout ce qu'il y a au dessus jusqu'au min
        max = data_month['P'].max()
        # idxmax = data_month['P'].drop_duplicates(keep='last').idxmax() # permet de trouver index de la dernière rep de cette valeur max
        idxmax = data_month['P'].idxmax()  # keep only the 1st val of the maxval to avoid invalid data of resuspension
        min1 = data_month['P'].loc[0:idxmax].min()  # min on the down profile
        idxmin = data_month['P'].idxmin()
        min2 = data_month['P'].loc[idxmax:].min()  # min on the up profile
        idxmin2 = data_month['P'].loc[idxmax:].idxmin()
        list_depth.append([min1, min2])
        print(min1, 'index_min', idxmin, '\n', max, idxmax)

        data_month2 = data_month.loc[idxmin:idxmax].copy()
        # data_month2 = data_month2[(data_month2 != 0).all(axis=1)]  # On enlève tous les 0

        # print(data_month2)
        # CRITERE 2 de selection des station qui fonctionnent ou non
        # if np.shape(data_month)[0]<15 : #si le fichier ou on a enlevé les 0 est trop court
        #     continue
        if np.shape(data_month2['P'].drop_duplicates().index.values)[
            0] < 10:  # Skip si le fichier ne contient pas plus de 10 profondeurs différentes
            continue
        # Done  : critere de selectio, : si longueur fichier trop petite, skipper
        print('shape data 1 : ', np.shape(data_month))
        print('shape data 2 : ', np.shape(data_month2), 'diff', np.shape(data_month)[0] - np.shape(data_month2)[0])
        # Je corrige P,   #Calculer les offsets  #

        if (min1 > 1.5 and month != 'June'):  # TODO A VOIR SI ON FAIT CA QUE POUR AOUT OU POUR TOUTES LES CAMPAGNES
            print('aha')
            data_month2['Pcorr'] = data_month2['P'].values - (
                    min1 - 0.5)  # on va considérer que première mesure est à 50cm sous surface
        else:
            print('Ohooh')
            data_month2['Pcorr'] = data_month2['P']

        ########################################################################################################
        # Traiter les données
        df_process = data_month2.iloc[:,
                     classe_first + 1:classe_last + 1]  # Je choisis les colonnes qui correspondent aux classes 3 à 30 inclu
        prof = data_month2['P']
        # print('prof',prof)
        # df_process=pd.concat([df_process,prof])
        print('df_process', df_process)

        df_SPMVC = pd.DataFrame()
        df_D50 = pd.DataFrame()
        df_SPMVC['SPMVC'] = df_process.sum(axis=1)
        df_SPMVC['P'] = data_month2['P'].values
        df_SPMVC['Pcorr'] = data_month2['Pcorr'].copy()
        # print(df_SPMVC, type(df_SPMVC))
        # print(list_width)

        for c in range(classe_first + 1, classe_last + 1):
            # boucle qui dépend de la taille de df_process, ie de 32 - classes extérieures virées, +1 car on utilise #1 et non indice 0
            classe = "#" + str(c)
            df_SPMVC["C vol " + classe] = df_process.iloc[:][classe] / list_width[c - 1]  # unité : µL L-1 µm-1
            # df_SPMVC["% VCcum "+classe]=df_process.iloc[:][classe]/df_SPMVC.iloc[:]['SPMVC'] #je calcule le SPMVC pour chaque classe
            # je calcule le C vol pour chaque classe
            if c - classe_first == 1:
                df_D50[classe] = df_SPMVC["C vol " + classe] * list_width[c - 1] / df_SPMVC.iloc[:]['SPMVC'] * 100
                # df_D50[classe] = df_SPMVC["% VCcum "+classe] * 100  # C vol cumulé : égal à C vol #1
            else:
                # df_D50[classe] = df_D50["#" + str(c - 1)] + (df_SPMVC["% VCcum "+classe]) * 100  # C vol cumulée : égal à C vol de la classe + le cumulé de la classe précédante
                df_D50[classe] = df_D50["#" + str(c - 1)] + df_SPMVC["C vol " + classe] * list_width[c - 1] / \
                                 df_SPMVC.iloc[:]['SPMVC'] * 100

            df_SPMVC["PSD " + classe] = df_SPMVC["C vol " + classe] / (
            df_SPMVC.iloc[:]['SPMVC'])  # c-1 car redevient un indice pour la list_width
            # C vol normalisée par la taille de la classe des particules
            # 22/04 Je rajoute le calcul du nombre de particules
            vol_sphere = (4 / 3) * np.pi * np.power((list_median_size[c - 1] / 2),
                                                    3) * 1000  # Vlumne en µm^3 qi pas *1000, en µL sinon
            df_SPMVC["n " + classe] = df_SPMVC[
                                          'C vol ' + classe] / vol_sphere  # nb de particule de la classe considérée
            # UNITE : (L-1 µm-1)

        # 24/04 : j'essaie de trouver le paramètre de Junge
        n_col = [col for col in df_SPMVC.columns if 'n ' in col]  # all col with number of part
        Junge = []
        l = list_median_size[classe_first:classe_last]
        for j in range(len(df_SPMVC[n_col])):
            D = df_SPMVC[n_col].iloc[j]
            Junge.append(Junge_parameter(l, D)[1])
        df_SPMVC['Junge'] = Junge

        # Calcul du D50 : trouver ou est le plus proche de 50 tout en étant inféreur
        print(df_D50)
        df_test = df_D50.T <= 50  # je sélectionne les valeurs inf ou égale à 50, renvoie tab boolean
        list_D50 = df_test[:].idxmin()  # je détecte la première valeur a false, donne l'index
        list_classe_D50, list_classe_D51 = [], []
        list_val_classe_D50, list_val_classe_D51 = [], []
        c = 0
        for i in range(len(list_D50)):
            # end=len(list_D50[i])
            ind_classe = int(list_D50.iloc[i][
                             1:3]) - 2  # J'extrai sous la forme d'int l'INDEX de la classe que l'on a trouvé avec list_D50 (numéro =num+1)
            list_classe_D50.append(
                ind_classe)  # list des valeurs des classes, à faire -1 pour aller chercher la valeur de la taille de la classe pour le calcul suivant
            # -1 car l'indice donne la première valeur à false, et non la dernière à true
            list_classe_D51.append(ind_classe + 1)
            val_classe = df_D50.iloc[c].iloc[
                ind_classe - classe_first]  # je selectionne ligne par ligne (c) la position détectée
            # de la colonne qui correspond à <50, numéro-val de la 1ere col qu'on a retiré, -1 car on est en index
            val_classe1 = df_D50.iloc[c].iloc[ind_classe + 1 - classe_first]
            list_val_classe_D50.append(val_classe)
            list_val_classe_D51.append(val_classe1)
            c = c + 1
        print('list D50 val classe', list_classe_D50)
        print('val classe', list_val_classe_D50)
        df_mediansize = pd.Series(list_median_size)

        reste_D50 = pd.DataFrame()  # list_val_classe_D50)
        reste_D50['val 50'] = list_val_classe_D50
        reste_D50['reste'] = 50 - reste_D50['val 50']
        reste_D50['val 51'] = list_val_classe_D51
        reste_D50['diff val'] = reste_D50['val 51'] - reste_D50['val 50']
        reste_D50['median class 50'] = df_mediansize[list_classe_D50].values
        reste_D50['median class 51'] = df_mediansize[list_classe_D51].values
        reste_D50['diff median'] = reste_D50['median class 51'] - reste_D50['median class 50']
        reste_D50['value finale'] = reste_D50['median class 50'] + reste_D50['reste'] * (
                reste_D50['diff median'] / reste_D50['diff val'])
        reste_D50['P'] = data_month2['P'].values
        print(reste_D50)

        # PAREIL MAIS POUR D90
        df_test2 = (df_D50.T) <= 90  # je sélectionne les valeurs inf ou égale à 50, renvoie tab boolean
        list_D90 = df_test2[:].idxmin()  # je détecte la première valeur a false, donne l'index
        list_classe_D90, list_classe_D91 = [], []
        list_val_classe_D90, list_val_classe_D91 = [], []
        c = 0
        for i in range(len(list_D90)):
            # end=len(list_D50[i])
            ind_classe = int(list_D90.iloc[i][
                             1:3]) - 2  # J'extrai sous la forme d'int l'INDEX de la classe que l'on a trouvé avec list_D50 (numéro =num+1)
            list_classe_D90.append(
                ind_classe)  # list des valeurs des classes, à faire -1 pour aller chercher la valeur de la taille de la classe pour le calcul suivant
            # -1 car l'indice donne la première valeur à false, et non la dernière à true
            list_classe_D91.append(ind_classe + 1)
            val_classe = df_D50.iloc[c].iloc[
                ind_classe - classe_first]  # je selectionne ligne par ligne (c) la position détectée
            # de la colonne qui correspond à <50, numéro-val de la 1ere col qu'on a retiré, -1 car on est en index
            val_classe1 = df_D50.iloc[c].iloc[ind_classe + 1 - classe_first]
            list_val_classe_D90.append(val_classe)
            list_val_classe_D91.append(val_classe1)
            c = c + 1

        reste_D90 = pd.DataFrame()  # list_val_classe_D50)
        reste_D90['val 90'] = list_val_classe_D90
        reste_D90['reste'] = 90 - reste_D90['val 90']
        reste_D90['val 91'] = list_val_classe_D91
        reste_D90['diff val'] = reste_D90['val 91'] - reste_D90['val 90']
        reste_D90['median class 90'] = df_mediansize[list_classe_D90].values
        reste_D90['median class 91'] = df_mediansize[list_classe_D91].values
        reste_D90['diff median'] = reste_D90['median class 91'] - reste_D90['median class 90']
        reste_D90['value finale'] = reste_D90['median class 90'] + reste_D90['reste'] * (
                reste_D90['diff median'] / reste_D90['diff val'])
        reste_D90['P'] = data_month2['P'].values
        print(reste_D90)

        df_SPMVC['D50'] = reste_D50['value finale'].values
        df_SPMVC['D90'] = reste_D90['value finale'].values
        df_SPMVC['Date'] = date  # data_month2['Date'].copy()
        # df_SPMVC['Hour'] = date #data_month2['Hour'].copy()

        if save:
            outfile = rep + 'Survey_' + month + '/' + station + '_SPM_#' + str(classe_first + 1) + '-#' + str(
                classe_last)
            # reste_D50.to_csv(outfile+'_D50', sep=' ', index=False, float_format='%.4f')
            df_SPMVC.to_csv(outfile, sep=' ', index=False, float_format='%.4f')
            print(outfile)
        # faire des moyennes ?
        # COMMENT choisir les couches ? Tout par mètres ? Sachant que problème d'offset.
        # Sur un mètre ? Checker la disparités via STD ?
        # df_process['P'] = prof  # J'ajoute une colonne au dataframe pour pouvoir faire des moyennes
        # moy_2m = df_process[df_process['P'] < 2]
        # print(moy_2m)

    else:
        # CHARGER fichier
        file_data = rep + 'Survey_' + month + '/' + station + '_SPM_#' + str(classe_first+1) + '-#' + str(classe_last)
        print('LOAD FILE ', file_data)
        if not os.path.exists(file_data):
            print('I skip ' + file_data + ' it does not exist')
            continue

        data = pd.read_csv(file_data, sep=' ')
        pmin = data['Pcorr'].min()
        pmax = data['Pcorr'].max()
        print('pmax', pmax)

        # MOYENNE par valeurs de P identiques
        print('AVERAGED THE VALUES OF IDENTICAL P')
        list_prof = list(set(data['Pcorr'].values))
        list_prof.sort()
        list_average_depth = []
        data_unique = pd.DataFrame()
        c = 0

        # hour = data['Hour'].loc[1]
        # date = data['Date'].loc[1]
        # date = datetime.strptime(date + ' ' + hour, '%Y-%m-%d %H:%M:%S')
        # to avoid error i drop the Date and Hour columns
        # date = calcul_date(data['int date'], data['int date 2'], int(year))
        # date = date[0]
        # data = data.drop(["Date", 'Hour'], axis=1)
        date = data['Date'][0]
        date = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
        print('date', date)

        for l in list_prof:
            # print('prof étudiée = ',l)
            # print(data['D50'].loc[data['Pcorr']==l])
            # print('STD sur D50', data['D50'].loc[data['Pcorr']==l].std(ddof=0))
            new_row = pd.DataFrame(data.loc[data['Pcorr'] == l].median())
            # print('new row ok')
            data_unique = pd.concat([data_unique, new_row.T], ignore_index=True)
            # print( 'concat ok ')
            data_unique.loc[c, 'STD D50'] = data['D50'].loc[data['Pcorr'] == l].std(
                ddof=0)  # donne la std de toutes les valeurs à la profondeur etudiée
            # print('add of STD val ok')
            c = c + 1
        print('ok for data_unique')

        # MOYENNE par couche #TODO : est ce qu'il vaut mieux faire ca a partir de data_unique (median) ou une moyenne de toutes les couches ?
        data_averaged = pd.DataFrame()  # QUESTION : tous les m ? 2m ? Divisé en 3 (sur,bot,mid ?)
        # Je teste en faisant un moyenne tous les metres
        p = 0
        if pmax <= 3:
            print('ATTENTION VAL MAX < 3m')  # , att aux moy de surf et bott')
        for i in range(int(pmax) + 1):
            condition = (data['Pcorr'] > p) & (data['Pcorr'] < p + 1)
            new_row = pd.DataFrame((data.loc[condition]).mean())
            data_averaged = pd.concat([data_averaged, new_row.T], ignore_index=True)
            data_averaged.loc[i, 'nb val'] = np.shape(data[condition])[0]
            data_averaged.loc[i, 'STD D50'] = data['D50'].loc[condition].std(
                ddof=0)  # donne la std de la moyenne de prof avec toutes les valeurs inclues
            p = p + 1

            # # Je sépare en 3 couches :
            # if pmax > 4:  # COndition , si non remplie, je ne fais qu'une couche
            #     nb_couche = 3
            #     list_couche = ['moy_couche' + str(i) for i in range(1, nb_couche + 1)]
            #     for i in range(0, nb_couche):
            #         condition = ((data_unique['Pcorr'] > i * pmax / 3) & (data_unique['Pcorr'] < (i + 1) * pmax / 3))
            #         list_couche[i] = data_unique[condition].copy()
            # else:
            #     nb_couche = 1
            #     list_couche = ['moy_couche' + str(i) for i in range(1, nb_couche + 1)]
            #     for i in range(0, nb_couche):
            #         condition = ((data_unique['Pcorr'] > i * pmax / 3) & (data_unique['Pcorr'] < (i + 1) * pmax / 3))
            #         list_couche[i] = data_unique[condition].copy()

        fontsize = 10
        outfile = rep + 'Survey_' + month + '/figure'
        if not os.path.exists(outfile + '/' + station):
            # if the demo_folder directory is not present
            # then create it.
            os.makedirs(outfile + '/' + station)
        outfile = outfile + '/' + station + '/'

        fixed = station.__contains__('F')

        ###################    figure de l'évolution de D50 % depth
        print('FIGURE of D50 evo')
        fig, ax = plt.subplots(ncols=1)
        if fixed and month == 'June':
            # fig.suptitle('D50 Fixed station ' + str(date.date()) + ' ' + str(date.time()))
            fig.suptitle('D50 Fixed station' + str(date.date()) + ' ' + str(date.time()))  # TODO change date format
        else:
            fig.suptitle('D50 Survey ' + month + ', ' + station)
        ax.scatter(data_unique['D50'], -data_unique['Pcorr'], marker='x', alpha=0.8, color='grey')
        ax.set_xlabel('D50', fontsize=fontsize)
        ax.set_ylabel('Depth (m)', fontsize=fontsize)
        ax.set_ylim(-15, 0)
        ax.set_xlim(0, 140)
        fig.savefig(outfile + station + '_D50', format='png')

        n = 1
        if n:
            ################## figure histogramme des PSD à chaque prof.
            PSD_col = [col for col in data_unique.columns if 'PSD' in col]
            data_PSD = data_unique[PSD_col]

            for d in range(np.shape(data_PSD)[0]):
                fig2, ax = plt.subplots()
                depth = str(data_unique['Pcorr'].loc[d])
                fig2.suptitle('Survey ' + month + ', ' + station + ' at depth=' + depth + ' m')
                # ax = data_PSD.plot.bar() #va ploter toutes les PSD pour chaque prof.
                ax = data_PSD.loc[d].plot.bar()  # va ploter toutes les PSD pour UNE prof.
                ax.set_xlabel('Median size (µm)', fontsize=fontsize)
                # ax.xaxis.set_minor_locator(MultipleLocator(5))
                ax.set_xticks(np.arange(1, len(list_median_size_name[2:30]) + 1, 1))
                ax.set_xticklabels(list_median_size_name[2:30], fontsize=fontsize - 2,
                                   rotation=30)  # TODO : a atuomatiser avec classe first classe last
                ax.set_ylabel('PSD ($µl^{-1}$ µm)', fontsize=fontsize)
                ax.set_ylim(0, (int(100 * data_PSD.loc[:].max().max()) + 1) / 100)
                fig2.savefig(outfile + station + '_rep_PSD_at' + depth, format='png')

            #########################  Figure averaged depth
            # nom_list_average_depth=['first', 'last', 'intermediate'] #on sait que rangé comme : first_m, last_m, middle_m1; middle metre 2 ...
            data_PSD_averaged = data_averaged[PSD_col]
            for i in range(np.shape(data_averaged)[0]):  # len(list_average_depth)):
                fig3, ax = plt.subplots()
                d = data_PSD_averaged.loc[i]
                # d = list_average_depth[l].iloc[:,3:-1:2]
                if fixed and month == 'June':
                    title = 'Survey ' + month + ', fixed station ' + str(date.time()) + ' Average around ' + str(
                        i + 1) + ' meter'
                else:
                    title = 'Survey ' + month + ', ' + station + ' Average around ' + str(i + 1) + ' meter'
                fig3.suptitle(title, fontsize=fontsize)
                # ax = data_PSD.plot.bar() #va ploter toutes les PSD pour chaque prof.
                ax = d.plot.bar()  # va ploter toutes les PSD pour UNE prof moyennée
                ax.set_xlabel('Median size (µm)', fontsize=fontsize)
                # ax.xaxis.set_minor_locator(MultipleLocator(5))
                ax.set_xticks(np.arange(0, len(list_median_size_name[2:30]), 1))
                ax.set_xticklabels(list_median_size_name[2:30], fontsize=fontsize - 2,
                                   rotation=30)  # TODO : a atuomatiser avec classe first classe last
                ax.set_ylabel('PSD ($µl^{-1}$ µm)', fontsize=fontsize)
                ax.set_ylim(0, (int(100 * data_PSD.loc[:].max().max()) + 1) / 100)
                fig3.savefig(outfile + station + '_PSD_averaged_' + str(i + 1) + 'm', format='png')

        if test_fill_surfer_file:

            # 1. lire l'heure et le jour depuis le fichier.
            # 2. aller chercher l'equivalent et la moyenne des X bins dans le fchier ADCP pour le plotter.
            rep_adcp = '/home/penicaud/Documents/Data/ADCP/Survey_' + month + '/'
            rep_CTD = '/home/penicaud/Documents/Data/CTD/Survey_' + month + '/'
            file_imer = rep_CTD + 'CTD_imer_' + month + '.xlsx'

            suffixe = '.csv'

            # 27/04/2023 Je créé un dictionnaire :
            dict = {'June': {'taille_bin': 0.3, 'offset': 0.6, 'sep': ' '},
                    'August': {'taille_bin': 0.1, 'offset': 0.66, 'sep': ','},
                    'Octobre': {'taille_bin': 0.1, 'offset': 0.66, 'sep': ','}}

            if month == 'June':
                ##############################   DONNEE ADCP #####################################################
                if date.day == 16:
                    file_adcp = rep_adcp + str(date.day) + '0' + str(date.month) + '_alldata_BT'
                elif date.day == 17:
                    if int(station[1:3]) <= 28:
                        file_adcp = rep_adcp + str(date.day) + '0' + str(date.month) + '_T3_alldata_BT'
                    else:
                        file_adcp = rep_adcp + str(date.day) + '0' + str(date.month) + '_T4_alldata_BT'
                elif date.day == 18:
                    file_adcp = rep_adcp + str(date.day) + '0' + str(date.month) + '_alldata_BT'

            elif month == 'August':
                int_sta = int(station[1:])
                if not fixed:
                    if int_sta <= 5:
                        # if transect == 'TA1':
                        hour_deb = '10h53'
                        hour_end = '12h24'
                    elif int_sta <= 12:  # transect == 'TA2':
                        hour_deb = '10h53'
                        hour_end = '12h24'
                        hour_deb = '12h41'
                        hour_end = '14h35'
                    elif int_sta <= 25:  # transect == 'SF1':
                        hour_deb = '14h36'
                        hour_end = '20h14'
                        # First values of current of the fixed station on the previous fil
                    elif int_sta <= 40:  # transect == 'TA3':
                        hour_deb = '15h01'
                        hour_end = '18h31'
                    elif int_sta <= 47:  # transect == 'TA4'
                        hour_deb = '06h46'
                        hour_end = '07h06'
                        # first minutes on another file
                else:
                    if station.__contains__('.'):  # in the case of not official fixed station
                        hour_deb = '14h36'
                        hour_end = '20h14'
                    else:
                        if int_sta <= 22:  # transect == 'SF_24.1':
                            hour_deb = '11h48'
                            hour_end = '19h49'
                        elif int_sta <= 38:  # transect == 'SF_24.2':
                            hour_deb = '19h50'
                            hour_end = '08h55'

                file_adcp = rep_adcp + str(date.day) + '0' + str(date.month) + '_' + hour_deb + '-' + hour_end + '_BT'

            elif month == 'Octobre':
                print('TODO')

            file_adcp = file_adcp + suffixe
            print('file_adcp', file_adcp)
            data_adcp = pd.read_csv(file_adcp, skiprows=11, low_memory=False,
                                    sep=dict[month]['sep'])  # , usecols=col_list)
            # Je cherche la ou correspond heure des stations
            # SI L'heure cherchée n'existe pas :
            data_profil = pd.DataFrame(
                (data_adcp.loc[(data_adcp['HH'] == date.hour) & (data_adcp['MM'] == date.minute)]).mean(
                    numeric_only=True))
            data_profil = data_profil.T
            adcp_not_good = data_profil.isnull().values.all()
            if adcp_not_good:  # Je check si toutes les valeurs sont à nan. Si oui, je le dis, et je sauterai la phase de figure
                print('File adcp do not contain the right time ')
            mag_cols = [col for col in data_profil.columns if
                        'Mag' in col]  # selection de seulement les colonnes voulues de Magnitude
            data_profil2 = data_profil[mag_cols].copy()
            # data_profil2=pd.DataFrame(data_profil2.dropna())
            depth_adcp = pd.DataFrame((float(data_profil2.columns[i][len(data_profil2.columns[i]) - 2:len(
                data_profil2.columns[i])])) * dict[month]['taille_bin'] + dict[month]['offset'] for i in
                                      range(np.shape(data_profil2)[1]))
            # ici j'ai construit la ligne des profondeurs à partir de taille_bin (dépend de chaque mois) et du numéro du bin concerné : 'Mag, mm/s, 3' = taille_bin*(3)+offset
            d_adcp = pd.DataFrame()
            d_adcp['depth'] = depth_adcp.copy()
            d_adcp['vitesse'] = data_profil2.T.values
            # d_adcp=d_adcp.rename(columns={0: 'vitesse'})
            # data_profil3 = signal.medfilt(data_profil2.T)  # filter 3
            # TODO resolve it with spatial filtering averaged ??
            list_depth_adcp = list(round(n, 1) for n in d_adcp['depth'].values)

            ############################   DONNEE DE SALINITÉ AVEC LA CTD IMER "#############################
            col_list_imer = ["Depth", "Temp", "Salinity", "Density", "Chl", "Turbidity"]
            # Done : skip si sheet=station n'existe pas
            from openpyxl import load_workbook

            wb = load_workbook(file_imer, read_only=True)  # open an Excel file and return a workbook

            if station in wb.sheetnames:
                print('sheet1 exists in imer file')
                data_imer = pd.read_excel(file_imer, station, skiprows=23,
                                          usecols=col_list_imer)  # lambda x : x > 0 and x <= 27 )#, usecols=col_list)
                print(data_imer)
                turb = signal.medfilt(data_imer['Turbidity'].values, 5)  # calcul de la turbidité avec filter median 5
                data_imer.insert(6, "Turbidity filtered 5", turb)  # insertion de la turb dans le df
            else:
                print('Data imer does not exist for this station : ', station)
                data_imer = pd.DataFrame(np.nan, index=range(len(list_depth_adcp)), columns=data_imer.columns)

            ############## CREATE a file with all values at same depth
            sal_extract = pd.DataFrame()
            D50_extract = pd.DataFrame()
            D90_extract = pd.DataFrame()
            for l in range(len(list_depth_adcp)):
                dmax = list_depth_adcp[l]
                if l != 0:
                    dmin = list_depth_adcp[l - 1]
                else:
                    dmin = 0
                # print('dmin et max', dmin, dmax) #depth imer : every 0.1m
                new_row = pd.DataFrame(
                    (data_imer.loc[(data_imer['Depth'] <= dmax) & (data_imer['Depth'] >= dmin)]).mean())
                # new_row = pd.DataFrame(data_imer.loc[(data_imer['Depth'] == list_depth_adcp)]) #extraire la valeur sans aucun traitement : pas de moyennage
                new_row_D50 = pd.DataFrame((data_unique.loc[
                    (data_unique['Pcorr'] <= dmax) & (data_unique['Pcorr'] >= dmin)]).mean())  # moyennage ?
                D50_extract = pd.concat([D50_extract, new_row_D50.T], ignore_index=True)
                sal_extract = pd.concat([sal_extract, new_row.T], ignore_index=True)
                # print('dmoy', sal_extract['Depth'][l])

            d_adcp['Salinity'] = sal_extract['Salinity'].values  # on ajoute les données de salinité
            d_adcp['Temp'] = sal_extract['Temp'].values  # on ajoute les données de salinité
            d_adcp['Turbidity filtered 5'] = sal_extract[
                'Turbidity filtered 5'].values  # on ajoute les données de salinité
            d_adcp['D50'] = D50_extract['D50'].values
            d_adcp['D90'] = D50_extract['D90'].values

            ############################### STATION RECAP ###########################"

            dist = data_station['Distance'].loc[(data_station['Stations'] == station)].values
            dist = np.ones(len(d_adcp['D90'])) * dist
            d_adcp['Distance'] = dist
            d_adcp['Time'] = len(d_adcp['D90']) * [date.time()]

            # Je sauvegarde le fichier, à utiliser pour toutes les données surfer.
            if save:
                path = rep_station + 'Recap_all_param_depth.xlsx'
                # CONTROLE SI FICHIER EXISTE :
                if not os.path.exists(path):
                    # create a new XLSX workbook
                    print('I create a new file')
                    import xlsxwriter

                    wb = xlsxwriter.Workbook(path)  # Create a workbook and add a worksheet.
                    worksheet = wb.add_worksheet(path)
                    wb.close()
                else:
                    # si la sheet n'existe pas, je la créé
                    wb = load_workbook(path, read_only=False)  # open an Excel file and return a workbook
                    if station in wb.sheetnames:
                        print('sheet1 exists in excel file')
                    else:
                        operations_sheet = wb.create_sheet(station)

                # wb = load_workbook(path, read_only=False)
                # wb.create_sheet(station)
                # d_adcp.to_excel(wb, sheet_name=station)
                # wb.save(path)
                # writer = pd.ExcelWriter(path, engine='xlsxwriter')
                # writer.close()
                # read the existing sheets so that openpyxl won't create a new one later
                book = load_workbook(path)
                writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                d_adcp.to_excel(writer, sheet_name=station, startrow=0, startcol=0, index=False)  # , header=header)
                writer.save()

            if figure:

                # Je sépare en 3 couches : #Done : a faire selon la vraie profondeur ==> On peut dire que c'est ok car on prend la prof max des 3 instruments.
                if pmax > 4 and not adcp_not_good:  # condition , si non remplie, je ne fais qu'une couche
                    option = 'adcp'  # 'adcp' ou ' moy' choisi si l'on prend la valeur max des 3 instruement ou si l'on prend la val de adcp
                    if option == 'moy':
                        max_depth = max(d_adcp['Salinity'].last_valid_index(), d_adcp['D50'].last_valid_index(),
                                        d_adcp['vitesse'].last_valid_index())
                        val_depth = d_adcp.loc[max_depth].loc[
                            'depth']  # c'est la pmax des 3 valeurs, issues de 3 instruments : ADCP, CTD imer, et LISST.
                    elif option == "adcp":
                        val_depth = d_adcp.loc[d_adcp['vitesse'].last_valid_index()].loc['depth']
                    nb_couche = 3
                    list_couche = ['moy_couche' + str(i) for i in range(1, nb_couche + 1)]
                    for i in range(0, nb_couche):
                        condition = (
                                (d_adcp['depth'] > i * val_depth / 3) & (d_adcp['depth'] < (i + 1) * val_depth / 3))
                        list_couche[i] = d_adcp[condition].copy()
                elif pmax <= 4:
                    nb_couche = 1
                    list_couche = ['moy_couche' + str(i) for i in range(1, nb_couche + 1)]
                    for i in range(0, nb_couche):
                        condition = (
                                (d_adcp['depth'] > i * val_depth / 3) & (d_adcp['depth'] < (i + 1) * val_depth / 3))
                        list_couche[i] = d_adcp[condition].copy()

                ###################    figure de l'évolution de D50 % depth AVEC LE PROFIL DE VITESSE
                fig, ax = plt.subplots(ncols=1)
                if fixed:
                    fig.suptitle('D50 Fixed station ' + str(date.date()) + ' ' + str(date.time()),
                                 fontsize=fontsize)
                else:
                    fig.suptitle('D50 Survey ' + month + ', ' + station, fontsize=fontsize)
                ax.scatter(data_unique['D50'], -data_unique['Pcorr'], marker='x', alpha=0.8,
                           color='grey')  # c=data_imer['Salinity'] ,cmap=cmap)#
                ax2 = ax.twiny()
                ax2.plot(d_adcp['vitesse'], -d_adcp['depth'], alpha=0.8,
                         color='orange')  # Ajouter la salinité en couleur ? cmap=data_imer['Salinity'] ne marche pas mais essayer
                ax2.set_ylabel('Current speed (mm/s)', fontsize=fontsize)
                ax.set_xlim(0, 140)
                ax2.set_xlim(100, 900)
                ax.set_ylim(-15, 0)
                ax.set_xlabel('D50', fontsize=fontsize)
                # ax.set_xlim(0,np.max())
                ax.set_ylabel('Depth (m)', fontsize=fontsize)
                fig.savefig(outfile + station + '_D50_fonction_vitesse', format='png')

                ###################    figure de l'évolution de D50 AVEC LE PROFIL DE VITESSE ET SALINITE
                if SspeedD50:
                    if adcp_not_good:
                        continue
                    list_marker = ['o', 'x', '^']
                    cmap = plt.cm.jet

                    vmin = 0
                    vmax = 160
                    # im= axSspeedD50.scatter(d_adcp['Salinity'], d_adcp['vitesse'] , c=c_plot, alpha=0.8, cmap=cmap, vmin=vmin, vmax=vmax)
                    im = axSspeedD50.scatter(list_couche[0]['Salinity'], list_couche[0]['vitesse'],
                                             marker=list_marker[0], c=list_couche[0]['D50'], alpha=0.8, cmap=cmap,
                                             vmin=vmin,
                                             vmax=vmax, label='layer 1')
                    if len(list_couche) > 1:
                        for i in range(1, len(list_couche)):
                            axSspeedD50.scatter(list_couche[i]['Salinity'], list_couche[i]['vitesse'],
                                                marker=list_marker[i], c=list_couche[i]['D50'],
                                                alpha=0.8, cmap=cmap, vmin=vmin,
                                                vmax=vmax, label='layer ' + str(i + 1))
                            if (station == 'S2'):  # TODO : ou premiere de aout ou octobre
                                axSspeedD50.legend(loc='upper right', fontsize=fontsize)

if SspeedD50:
    cbar = plt.colorbar(im, label='D50 (µm)', ax=axSspeedD50)  # , ticks=1)
    figSspeedD50.savefig(rep + 'Survey_' + month + '/figure/Diagramm_Speed_Sal_D50_couche', format='png')

print('end    ')
