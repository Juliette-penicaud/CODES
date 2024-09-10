# 17/04/23 : j'ai coupé la partie ou je créée les fichiers excel (pas particulièrement pour surfer)
# 27/05 : DONE : tableaux pour aout et octobre avec les Pcorr du LISST v2705 avec les 0 pour avoir prof max et
# comp avec ADCP et CTD pour aviser ensuite si juste offset_flottant ou pire
# 05/06 : Done: refaire le tableau dès 0m de prof, i.e NE PAS caler les profs selon ADCP,
# 06/06/23 : Now all the stations are referenced and not only the one in the LISST_sheets (all the one in data_station)
# 08/06/23 : Tab from 0m and not from the first bin of data_recap (0.66). Equal interval from surface to bottom
# TODO 07/06/23 : I want to add the CTD LOG to have the salinity in the depth

import pandas as pd
import numpy as np
import sys, os
from datetime import datetime
import scipy.signal as signal
import re
from openpyxl import load_workbook
import gsw
import glob
import scipy.integrate as integrate

# VARIABLE AND PARAMETER
year = '2022'
list_month = ['June', 'August', 'Octobre']
i = 2  # 0 1 2
month = list_month[i]

save = False
depth_register = False
# option = 'CTD'  # option to set the "base" instrument

# LISST FILE : sert juste pour les stations
rep_LISST = '/home/penicaud/Documents/Data/LISST/'
suffixe = '.xlsx'
file_LISST = rep_LISST + 'LISST_' + month + '_' + year + suffixe
print('file LISST : ', file_LISST)

# STATIONS DATA
rep_station = '/home/penicaud/Documents/Data/Survey_' + month + '/'
file_station = rep_station + 'Stations_' + month + '.xlsx'
print('file stations recap', file_station)
data_station = pd.read_excel(file_station)  # , usecols=['Stations', 'Time', 'Distance','% '])

# ADCP FILE
rep_adcp = '/home/penicaud/Documents/Data/ADCP/Survey_' + month + '/'

# CTD IMER file
rep_CTD = '/home/penicaud/Documents/Data/CTD/Survey_' + month + '/'
file_imer = rep_CTD + 'CTD_imer_' + month + '.xlsx'
print('file imer : ', file_imer)

# CTD LOG
rep_LOG = rep_CTD + 'CTD_LOG/Stations/'

# Dictionary of the transects and their last stations number
dict_transect = {'June': {'T1': 10, 'T2': 17, 'T3': 28, 'T4': 39},
                 'August': {'T1': 5, 'T2': 12, 'T3': 27, 'T4': 34},
                 'Octobre': {'T1': 7, 'T2': 21, 'T3': 29, 'T4': 35}}

list_prof_CTD, list_prof_ADCP, list_prof_LOG = [], [], []
list_sta_CTD, list_sta_ADCP, list_sta_LOG = [], [], []

dict_month = {'June': {'nrows': 87, 'char': 'S', 'depth_max': 20, 'step_depth': 0.3},
              'August': {'nrows': 95, 'char': 'A', 'depth_max': 20, 'step_depth': 0.1},
              'Octobre': {'nrows': 111, 'char': 'O', 'depth_max': 20, 'step_depth': 0.1}}

char = dict_month[month]['char']
rep = '/home/penicaud/Documents/Data/Survey_' + month
file_station = rep + '/Stations_' + month + '.xlsx'
df_global = pd.read_excel(file_station, sheet_name=0, nrows=dict_month[month]['nrows'])  # read the stations name
df_global = df_global.dropna(subset=['Stations'])
list_station = df_global['Stations'].values  # 07/06/23

# # creation of a list from the LISST stations
# d1 = pd.ExcelFile(file_LISST)  # LISST df
# list_sheet_LISST = d1.sheet_names  # list of the file_LISST sheets
# list_sheet_LISST = [col for col in list_sheet_LISST if char in col]
# list_sheet_LISST_station = [col for col in list_sheet_LISST if not 'F' in col]
# list_sheet_LISST_f = [col for col in list_sheet_LISST if 'F' in col]
# # SAME for CTD IMER
# d2 = pd.ExcelFile(file_imer)  # CTDlist_sheet_IMER df
# list_sheet_IMER = d2.sheet_names  # list of the CTD file
# list_sheet_IMER = [col for col in list_sheet_IMER if char in col]
# list_sheet_IMER_station = [col for col in list_sheet_IMER if not 'F' in col]
# list_sheet_IMER_f = [col for col in list_sheet_IMER if 'F' in col]
#
# # From 2 lists : creation of a list with all stations (and unique) :
# list_sheet_f = list_sheet_LISST_f.copy()
# list_sheet_station = list_sheet_LISST_station.copy()
# for i in range(len(list_sheet_IMER_station)):
#     if list_sheet_IMER_station[i] not in list_sheet_station:
#         list_sheet_station.append(list_sheet_IMER_station[i])
# list_sheet_station = sorted(list_sheet_station, key=lambda s: int(re.search(r'\d+', s).group()))
#
# for i in range(len(list_sheet_IMER_f)):
#     if list_sheet_IMER_f[i] not in list_sheet_f:
#         list_sheet_f.append(list_sheet_IMER_f[i])
# list_sheet_f = sorted(list_sheet_f, key=lambda s: int(re.search(r'\d+', s).group()))
#
# list_station = list_sheet_station + list_sheet_f

# Record the lowest depth
list_depth = []
print('list_station', list_station)


# Definition fonction pour csv
def check_sta_autre_csv(station, file, col_list):
    # si jamais une des stations n'existe pas ou alors si une des stations est invalide
    list_test_sta = [file, file + '-1', file + '-2', file + '-3', file + '-4']
    df_test = pd.DataFrame(np.nan, index=range(2), columns=col_list)
    for sta in list_test_sta:
        if os.path.exists(sta):
            print(station + ' doesnt exist but ' + sta + ' exists in LISST file')
            df_test2 = pd.read_csv(sta, sep=' ', usecols=col_list)
            if np.shape(df_test2)[0] > np.shape(df_test)[0]:  # je prends le fichier le plus long de toutes les possib
                df_test = df_test2.copy()
        if np.shape(df_test.shape)[0] < 3:  # cad aucun de la liste ne convient.
            df_final = pd.DataFrame(np.nan, index=range(2), columns=col_list)  # 2 arbitraire
        else:
            df_final = df_test.copy()
    return df_final


# Definition fonction pour xls
def check_sta_autre_xls(station, wb, file, skiprow, col_list):
    # si jamais une des stations n'existe pas ou alors si une des stations est invalide
    list_test_sta = [station, station + '-1', station + '-2', station + '-3', station + '-4']
    df_test = pd.DataFrame(np.nan, index=range(2), columns=col_list)
    for sta in list_test_sta:
        if sta in wb.sheetnames:
            print(station + ' doesnt exist but ' + sta + ' exists in imer file')
            df_test2 = pd.read_excel(file, sta, skiprows=skiprow, usecols=col_list)
            if np.shape(df_test2)[0] > np.shape(df_test)[0]:  # je prends le fichier le plus long de toutes les possib
                df_test = df_test2.copy()
        if np.shape(df_test.shape)[0] < 3:  # cad aucun de la liste ne convient.
            df_final = pd.DataFrame(np.nan, index=range(len(list_depth_adcp)),
                                    columns=col_list)  # columns=data_imer.columns)
        else:
            df_final = df_test.copy()
    return df_final


def check_best_data():  # TODO
    # something to check if it is the best or if another one is better station-1 ect.
    print("TODO ")


def calcul_N2(T, S, depth):
    # Calcul de N2
    p = 0  # 1013.25
    p = depth # 13/02/24 : test with p!=0
    SA = gsw.SA_from_SP(S, p, 106.5, 20.5)
    CT = gsw.CT_from_t(SA, T.values, p)
    PT = gsw.pt0_from_t(SA, T.values, p)
    p_n2 = np.arange(0, len(T)/10, 0.1, dtype=float)
    #p_n2 = np.arange(0, len(T) * 0.001013, 0.001013, dtype=float)  # Gives only even number, keep len(T) because it will
    # give +1 if odd nb
    # tableau qui va de 0 à la surface à la pression de profondeur donnée par len(temperature) en dm (car avec la ctd imer
    # une mesure tous les 0.1m, len turbidity donne le nombre de mesures si 54 = 5.4m de prof) *0.001013 dbar car on
    # perd un atm tous les 10m : 0.001013 db tous les 0.1m
    # p_n2=p*p_n2
    [N2, p_mid] = gsw.Nsquared(SA, CT, p_n2[0:len(SA)], 20.5)  # 08/06 : change p_n2 to p_n2[0:len(SA)] to have the
    # right lenght, i.e the same as SA and CT, in case of odd number.
    # ATTENTION §§§ len(p_mid) et len(N2)= len de tous le reste -1 ==> PQ ? Rajouter pour éviter décalage
    N2_bis = np.ones(len(SA), dtype=float)  # A savoir pourquoi : len(N2) = len(SA)-1
    N2_bis[0:len(N2_bis) - 1] = N2_bis[0:len(N2_bis) - 1] * N2  # On remet à la taille
    N2_bis[len(N2_bis) - 1] = np.nan  # BIDOUILLE pour ajoueter un NAN à la fin
    N2_bis = pd.DataFrame(N2_bis, columns=['N2'])
    return N2_bis


def calcul_nb_Ri(nor, eas, density, taille_bin):
    g = 9.81
    du = np.gradient(nor, taille_bin)
    dv = np.gradient(eas, taille_bin)
    dD = np.gradient(density, taille_bin)
    Ri = -g * dD / (density * (np.square(du) + np.square(dv)))
    return Ri


def calcul_excess_density_methodA(data_recap):
    # ATTENTION : first valid index suppose que l'on a toujours la 1ere val<1m pour TOUS instrum
    rho_o = 1100  # density organic matter kg/m3
    rho_s = 2650  # density mineral matter in flocs
    rho = data_recap['Density'].iloc[data_recap['Density'].first_valid_index()]  # unit : kg/m3
    omega = (100 - data_recap['% OM'].iloc[data_recap['% OM'].first_valid_index()]) / 100
    SPM = data_recap['Concentration GFF surf'].iloc[
        data_recap['Concentration GFF surf'].first_valid_index()]  # unit : mg/L g/m3
    SPMVC = data_recap['SPMVC'].iloc[data_recap['SPMVC'].first_valid_index()]  # unit : µL/L
    rho_p = omega * rho_s + (1 - omega) * rho_o
    Delta_rhof = SPM / SPMVC * (1 - rho / rho_p)  # unit : g/mL
    return Delta_rhof * 1000  # *1000 to have in g/L


def calcul_excess_density_methodB(data_recap):
    rho_o = 1100  # density organic matter
    rho_s = 2650  # kg/m^3 density mineral matter in flocs
    df = 2  #
    rho = data_recap['Density']
    D50 = data_recap['D50']
    omega = (100 - data_recap['% OM']) / 100
    Dp = 4  # unit : µm
    rho_p = np.ones(len(rho)) * (omega * rho_s + (1 - omega) * rho_o)
    Delta_rhof = (rho_p - rho) * (Dp / D50) ** (3 - df)
    return Delta_rhof  # unit :


def calcul_simpson_primitive(data_recap):
    # 1/H*(intgz(rhomoy-rho)dz) sur H
    if type(data_recap['Salinity'].last_valid_index()) != int:
        I = np.nan
    else:
        max_depth = data_recap['Depth'].iloc[data_recap['Salinity'].last_valid_index() - 1]
        g = 9.81
        rho_moy = np.mean(data_recap['Density'])
        rho = data_recap['Density'][data_recap['Depth'] == max_depth]

        I = - (1 / 2) * g * max_depth * (rho_moy - rho)
    return I


def calcul_ws(data_recap):
    mu = 1.002 * 10 ** -3  # dynamic viscosity of water to be in kg/(m s)
    Delta_rhof = data_recap['Delta rhof methodB']
    D50 = data_recap['D50']
    g = 9.81
    df = 2
    # ws = 1/(18*mu) * Delta_rhof * g * Dp**(3-df) * D50**(df-1) * 10**-6 # *10**-6 to have the result in mm/s
    ws = 1 / (18 * mu) * Delta_rhof * g * D50 ** 2 * 10 ** -9  # *10**-6 to have the result in mm/s
    return ws


def calcul_df(data_recap):
    Delta_rhof = data_recap["Delta rhof methodA"].iloc[data_recap['Delta rhof methodA'].first_valid_index()]
    rho_o = 1100  # density organic matter
    rho_s = 2650  # kg/m^3 density mineral matter in flocs
    rho = data_recap['Density'].iloc[data_recap['Density'].first_valid_index()]
    D50 = data_recap['D50'].iloc[data_recap['D50'].first_valid_index()]
    omega = (100 - data_recap['% OM']) / 100
    Dp = 4  # unit : µm
    rho_p = (omega * rho_s + (1 - omega) * rho_o)
    df = 3 - np.log(Delta_rhof / (rho_p - rho)) / np.log(Dp / D50)
    return df


def calcul_simpson(data_recap):
    if data_recap['Salinity' ].isnull().all():
        I = np.nan
    else:
        max_depth = data_recap['Depth'].iloc[data_recap['Salinity'].last_valid_index() - 1]
        g = 9.81
        rho_moy = np.nanmean(data_recap['Density'])
        nb = data_recap['Density'].count()
        I = 0
        depth_old = 0
        for i in range(nb - 1):
            depth = data_recap['Depth'].iloc[i]
            dz = depth - depth_old
            rho = data_recap['Density'].iloc[i]
            val = depth * (rho_moy - rho) * dz
            I = I + val
            depth_old = depth
        I = - I * g / max_depth  # minus ??
    return I

    # def f_integrate(x,r,rmoy):
    #     g = 9.81
    #     return g*x(rmoy-r)
    # g=9.81
    # rho_moy=np.mean(data_recap['Density'])
    # #divisé par depth ou max depth ? Par depth mais ne représente pas forcement toute la colonne.
    # #max_depth = max(data_recap['Salinity'].last_valid_index(), data_recap['D50'].last_valid_index(), data_recap['vitesse'])
    # max_depth=data_recap['Depth'].iloc[data_recap['Salinity'].last_valid_index()-1]
    # val=1/max_depth * integrate.quad(g*z(rho_moy-data_recap['Density']))
    # r=data_recap['Density'][data_recap['Depth']<=max_depth]
    # quad(f_integrate, -max_depth, 0, args=(r, rho_moy))


def calcul_G(data_recap):
    dmax = data_recap["D90"] * 10 ** -6
    v = 1.002 * 10 ** -3  # kg / (m s)
    G = v / dmax ** 2
    return G


print('*********************************************************', '\n')
list_station = list_station[52:60]
for station in list_station:  # len(indice)] : #indice #45:46
    print('station', station)
    fixed = station.__contains__('F')

    if station.__contains__('-'):
        print('Je suis dans le cas dune double station, je regarde donc la station officielle précédante')
        station_init = station  # on checkera si station_init==station
        station = station[0:len(station) - 2]
        # station devient juste le radical, sans le '-1', plus pratique pour faire toutes les manips sans if
    else:
        station_init = station

    # Determiner la liste des transect :
    if fixed:
        if station.__contains__('.'):
            transect = 'small fixe'
        else:
            transect = 'fixe'
    else:
        for t in ['T1', 'T2', 'T3', 'T4']:
            if int(station[1:]) <= dict_transect[month][t]:
                transect = t
                break  # necessaire car condtion ne fixe pas la borne inf

    # Determiner heure station à partir du doc recap stations
    date = data_station['Time'].loc[(data_station['Stations'] == station)].values
    date = pd.to_datetime(date[0])

    ##################" LOAD LISST file ###########################################"
    classe_first = '2' # 2 if I wanna delete only one class, 3 if I want to not consider the 2 first classes
    classe_last = '31'
    file_LISST = rep_LISST + 'Survey_' + month + '/' + station + '_SPM_#' + classe_first + '-#' + classe_last
    print('open file LISST', file_LISST)
    if not os.path.exists(file_LISST):  # If no LISST file, data_LISST with NAN values
        print('file LISST does not exist')
        list_col_LISST = ['Pcorr', 'D50', 'D90', 'SPMVC', 'Junge']
        data_LISST = check_sta_autre_csv(station, file_LISST, list_col_LISST)
        # data_LISST = pd.DataFrame(np.nan, index=range(2), columns=)
        data_unique = data_LISST.copy()
        # print('data LISST', data_LISST)
    else:
        print('File LISST exists')
        data_LISST = pd.read_csv(file_LISST, sep=' ')
        # On enregistre la date et heure
        # hour_LISST = data_LISST['Hour'].loc[1]
        # date_LISST = data_LISST['Date'].loc[1]
        # date_LISST = datetime.strptime(date_LISST + ' ' + hour_LISST, '%Y-%m-%d %H:%M:%S')
        date_LISST = datetime.strptime(data_LISST['Date'][0], '%Y-%m-%d %H:%M:%S')
        # to avoid error i drop the Date and Hour columns
        data_LISST = data_LISST.drop(["Date"], axis=1)

        # CALCUL SUR LES DONNEES LISST
        pmin = data_LISST['Pcorr'].min()
        pmax = data_LISST['Pcorr'].max()
        print('pmax', pmax)
        # I NEED TO RECORD PMAX AND THEN TO REMOVE ALL 0.0 SO THEY DO NOT HAVE IMPACT ON AVERAGES
        # MOYENNE par valeurs de P identiques
        list_prof = list(set(data_LISST['Pcorr'].values))
        list_prof.sort()
        list_average_depth = []
        data_unique = pd.DataFrame()
        c = 0

        # data_month2 = data_month2[(data_month2 != 0).all(axis=1)]  # On enlève tous les 0
        # As I deleted the values of the
        # if not (len(data_LISST[(data_LISST == 0).all(axis=1)]) == 0):
        for l in list_prof:  # for all the different depth recorded in the file, keep the median value of each depth
            # WARNING : derive a median value from all the parameter and DO NOT keep the median value : i.e :
            # sum of all PSD do not lead to SPMVC
            # new_row.T[['n #' + str(i) for i in range(2,31)]]
            new_row = pd.DataFrame(data_LISST.loc[data_LISST['Pcorr'] == l].median())
            # new_row = data_LISST.loc[data_LISST['Pcorr'] == l].median().copy()
            data_unique = pd.concat([data_unique, new_row.T], ignore_index=True)
            data_unique.loc[c, 'STD D50'] = data_LISST['D50'].loc[data_LISST['Pcorr'] == l].std(
                ddof=0)  # donne la std de toutes les valeurs à la profondeur etudiée
            c = c + 1
        print('ok for data_unique LISST')

        # MOYENNE par couche #TODO : est ce qu'il vaut mieux faire ca a partir de data_unique (median) ou une moyenne de toutes les couches ?
        data_averaged = pd.DataFrame()  # QUESTION : tous les m ? 2m ? Divisé en 3 (sur,bot,mid ?)
        # Je teste en faisant un moyenne tous les metres ==> Not useful
        p = 0
        if pmax <= 3:
            print('ATTENTION VAL MAX < 3m')  # , att aux moy de surf et bott')
        for i in range(int(pmax) + 1):
            condition = (data_LISST['Pcorr'] > p) & (data_LISST['Pcorr'] < p + 1)
            new_row = pd.DataFrame((data_LISST.loc[condition]).mean())
            # TODO : quoi faire ? Avec median i.e data unique ou mean i.e avec data ?
            data_averaged = pd.concat([data_averaged, new_row.T], ignore_index=True)
            data_averaged.loc[i, 'nb val'] = np.shape(data_LISST[condition])[0]
            data_averaged.loc[i, 'STD D50'] = data_LISST['D50'].loc[condition].std(
                ddof=0)  # donne la std de la moyenne de prof avec toutes les valeurs inclues
            p = p + 1

    dict_adcp = {'June': {'taille_bin': 0.3, 'offset': 0.6, 'sep': ' ', 'turbTOconcentration': 0.73},
                 'August': {'taille_bin': 0.1, 'offset': 0.6, 'sep': ',', 'turbTOconcentration': 1.09},
                 'Octobre': {'taille_bin': 0.1, 'offset': 0.6, 'sep': ',', 'turbTOconcentration': 0.95}}
    # normally, offset = 0.66, set to 0.7
    # turbTOconcentration d'aout et octobre =  facteur GFF+Nucle

    #################      ADCP DATA   ####################################
    if month == 'June':
        if date.day == 17:
            if int(station[1:3]) <= 28:
                term = '_T3'
            else:
                term = '_T4'
        else:
            term = ''
        skiprow = 11
        file_adcp = rep_adcp + str(date.day) + '0' + str(date.month) + term + '_alldata_BT.csv'

    elif month == 'August':
        if not fixed:
            int_sta = int(station[1:])
            if int_sta <= 5:
                # if transect == 'TA1':
                hour_deb = '10h53'
                hour_end = '12h24'
            elif int_sta <= 12:  # transect == 'TA2':
                hour_deb = '10h53'
                hour_end = '12h24'
                hour_deb = '12h41'
                hour_end = '14h35'
            elif int_sta <= 27:  # transect == 'TA3':
                hour_deb = '15h01'
                hour_end = '18h31'
            elif int_sta <= 34:  # transect == 'TA4':
                hour_deb = '06h46'
                hour_end = '07h06'
                # first minutes on another file
        else:
            if station.__contains__('.'):  # in the case of not official fixed station
                hour_deb = '14h36'
                hour_end = '20h14'
            else:
                int_sta = int(station[2:])
                if int_sta <= 22:  # transect == 'SF_24.1':
                    hour_deb = '11h48'
                    hour_end = '19h49'
                elif int_sta <= 38:  # transect == 'SF_24.2':
                    hour_deb = '19h50'
                    hour_end = '08h55'

        if date.day == 13:
            day = 12
        else:
            day = date.day

        file_adcp = rep_adcp + str(day) + '0' + str(date.month) + '_' + hour_deb + '-' + hour_end + '_BT.csv'
        skiprow = 11

    elif month == 'Octobre':
        if date.day == 2:
            hour_deb = '17h19'
            hour_end = '18h28'
            term = '_T1_'
            skiprow = 11
        elif date.day == 3:
            hour_deb = '10h20'
            hour_end = '18h00'
            term = '_T2_T3_'
            skiprow = 11
        elif date.day == 4 and not fixed:
            hour_deb = '07h40'
            hour_end = '08h24'
            term = '_T4_'
            skiprow = 11
        elif (date.day == 4 or date.day == 5) and fixed:
            hour_deb = ''
            hour_end = ''
            term = '_FO'
            skiprow = 11
        else:
            print('PROBLEM TRANSECT octobre')

        if date.day == 5 and fixed:
            day = 4  # bidouille because all is in the same file
        else:
            day = date.day

        file_adcp = rep_adcp + '100' + str(day) + term + hour_deb + '_' + hour_end + '.csv'

    #######################        ADCP       #########################################""
    print('file_adcp', file_adcp)
    data_adcp = pd.read_csv(file_adcp, skiprows=skiprow, low_memory=False,
                            sep=dict_adcp[month]['sep'])  # , usecols=col_list)

    # Je cherche la ou correspond heure des stations
    delais = 1  # Todo : moyenner sur plus que une min ? ATTENTION nb profil différent d'une campagne à l'autre
    cond = (data_adcp.loc[(data_adcp['HH'] == date.hour) & (data_adcp['MM'] >= date.minute)
                          & (data_adcp['MM'] <= date.minute + delais) & (data_adcp['DA'] == date.day)])
    # 24/11/23 I add the condition on the day fo octobre fixe station ATTENTION LES PROFILS SONT MOYENNES SUR 2 MN
    data_profil = pd.DataFrame(cond.mean(numeric_only=True))
    data_profil = data_profil.T

    # SI L'heure cherchée n'existe pas :
    adcp_not_good = data_profil.isnull().values.all()
    if adcp_not_good:  # Je check si toutes les valeurs sont à nan. Si oui, je le dis, et je sauterai la phase de figure
        print('File adcp does not contain the right time ')

    mag_cols = [col for col in data_profil.columns if
                'Mag' in col]  # selection de seulement les colonnes voulues de Magnitude
    eas_cols = [col for col in data_profil.columns if
                'Eas' in col]
    nor_cols = [col for col in data_profil.columns if
                'Nor' in col]
    ver_cols = [col for col in data_profil.columns if
                'Ver' in col]
    data_profil2 = data_profil[mag_cols].copy()
    # data_profil2=pd.DataFrame(data_profil2.dropna())
    depth_adcp = pd.DataFrame((float(
        data_profil2.columns[i][11:len(data_profil2.columns[i])])) *
                              dict_adcp[month]['taille_bin'] + dict_adcp[month]['offset']
                              for i in range(np.shape(data_profil2)[1]))
    # 30/05/2023 : modif of len(data_profil2.columns[i])-2 to 11 TO have the number > 99 for August and Octobre survey
    # ici j'ai construit la ligne des profondeurs à partir de taille_bin (dépend de chaque mois) et du numéro
    # du bin concerné : 'Mag, mm/s, 3' = taille_bin*(3)+offset

    data_recap = pd.DataFrame()
    data_recap['Depth'] = depth_adcp.copy()

    #####################    ADD THE VALUES OF ADCP ############
    nor = data_profil[nor_cols].T.values
    eas = data_profil[eas_cols].T.values
    data_recap['vitesse mag'] = data_profil2.T.values
    data_recap['vitesse east'] = eas
    data_recap['vitesse north'] = nor
    data_recap['module vitesse u'] = np.sqrt(np.square(eas) + np.square(nor))
    data_recap['grad vitesse horiz'] = np.gradient(data_recap['module vitesse u'], dict_adcp[month]['taille_bin'])
    data_recap['angle'] = np.rad2deg(np.arctan(nor / eas)) # abs of nor and east ?

    # TRICK TO ADD ROWS AT BEGINNING TO HAVE REGULAR STEP FROM 0
    # nb_row to add : int(depth_adcp.loc[0]/dict_month[month]['step_depth']) with nan value BUT depth.
    list_val = np.zeros(len(data_recap.columns))  # +1 for the indexer
    list_val[:] = np.nan
    index = 0
    val_row = round((depth_adcp.loc[0].values / dict_month[month]['step_depth'])[0])
    data_recap = data_recap.set_index(data_recap.index + val_row)
    for nb_row in range(val_row):
        val_depth = nb_row * dict_month[month]['step_depth']
        list_val[0] = val_depth
        new_row_adcp = pd.DataFrame(list_val)
        new_row_adcp = new_row_adcp.T
        new_row_adcp = new_row_adcp.rename(columns={i: data_recap.columns[i] for i in range(len(data_recap.columns))})
        new_row_adcp = new_row_adcp.set_index([pd.Index([index])])
        data_recap = pd.concat([data_recap, new_row_adcp], ignore_index=False)
        index = index + 1
    data_recap = data_recap.sort_index()

    if depth_register:
        if np.all(np.isnan(data_recap['vitesse mag'])):  # if there is only NAN values
            print('skip cause all ADCP is at Nan')
            # continue
        else:
            if len(a_ADCP) == 0:
                list_prof_ADCP.append(data_recap['Depth'].iloc[data_recap['vitesse mag'].last_valid_index()])
                list_sta_ADCP.append(station)
            elif len(list_sta_ADCP) > 0:
                if list_sta_ADCP[-1] == station:  # if previous recorded station is the same, we skip
                    continue
                else:
                    list_prof_ADCP.append(data_recap['Depth'].iloc[data_recap['vitesse mag'].last_valid_index()])
                    list_sta_ADCP.append(station)

    # data_recap['grad vitesse mag'] = np.gradient(data_recap['vitesse mag'].values, taille_bin)  # approximation : que vitesse
    # horizontale et pas verticale
    # data_recap=data_recap.rename(columns={0: 'vitesse'})
    # data_profil3 = signal.medfilt(data_profil2.T)  # filter 3
    # TODO resolve it with spatial filtering averaged ??
    list_depth_adcp = list(round(n, 1) for n in data_recap['Depth'].values)

    ############################### STATION RECAP ###########################"

    dist = data_station['Distance'].loc[(data_station['Stations'] == station)].values
    dist = np.ones(len(data_recap['Depth'])) * dist
    data_recap['Distance'] = dist
    data_recap['Time'] = len(data_recap['Depth']) * [date]

    OM = data_station['% OM'].loc[(data_station['Stations'] == station)].values
    Turb_surf = data_station['Turbidity'].loc[(data_station['Stations'] == station)].values
    Tide_situation = data_station['Percentage of tide'].loc[(data_station['Stations'] == station)].values
    Tide = data_station['Tide'].loc[(data_station['Stations'] == station)].values
    GFF = data_station['GFF'].loc[(data_station['Stations'] == station)].values
    Nucle = data_station['NUCLEPORE'].loc[(data_station['Stations'] == station)].values

    print('ok for sta parameter ')
    data_recap['Transect'] = transect
    data_recap['Percentage of tide'] = np.ones(len(data_recap['Depth'])) * float(Tide_situation)
    data_recap['Tide level'] = np.ones(len(data_recap['Depth'])) * float(Tide)
    data_recap['% OM'] = np.ones(len(data_recap['Depth'])) * float(OM)
    data_recap['Concentration GFF surf'] = np.ones(len(data_recap['Depth'])) * float(GFF)
    data_recap['Concentration Nucle surf'] = np.ones(len(data_recap['Depth'])) * float(Nucle)
    data_recap['Turb surf'] = np.ones(len(data_recap['Depth'])) * float(Turb_surf)
    print('ok for turbidity surface')

    ############################      LOG CTD           ##############################
    f_LOG = glob.glob(rep_LOG + station + '_' + '*')
    print('f_LOG', f_LOG)
    col_list_LOG_init = ["Depth (Meter)", "Temperature (Celsius)", "Conductivity (MicroSiemens per Centimeter)",
                         "Salinity (Practical Salinity Scale)", "Density (Kilograms per Cubic Meter)"]
    col_list_LOG = ['Depth', 'Temperature', 'Conductivity', 'Salinity', 'Density']
    if len(f_LOG) == 0:
        print('No LOG file corresponding')
        LOG_extract = pd.DataFrame(np.nan, index=range(len(list_depth_adcp)), columns=col_list_LOG)
    else:
        f_LOG = f_LOG[0]
        df_LOG = pd.read_csv(f_LOG, skiprows=28,
                             usecols=col_list_LOG_init)  # lambda x : x > 0 and x <= 27 )#, usecols=col_list)
        df_LOG = df_LOG.rename(columns={col_list_LOG_init[i]: col_list_LOG[i] for i in range(len(col_list_LOG))})
        # hour_LOG = str(int(f_LOG[-10:-8]) + 7)  # To have the local time UTC+7
        # mn_LOG = f_LOG[-8:-6]  # minutes
        # sec_LOG = f_LOG[-6:-4]
        # if int(hour_LOG) >= 24:
        #     day = str(int(f_LOG[-13:-11]) + 1)
        #     hour_LOG = str(int(hour_LOG) % 24)
        #     print('CASE hour>24, new hour ', hour_LOG)
        # else:
        #     day = f_LOG[-13:-11]
        # print("hour LOG", hour_LOG, mn_LOG)
        LOG_extract = pd.DataFrame()
        for l in range(len(list_depth_adcp)):
            dmax = list_depth_adcp[l]
            if l != 0:
                dmin = list_depth_adcp[l - 1]
            else:
                dmin = 0
            # print('dmin et max', dmin, dmax) #depth imer : every 0.1m
            new_row_LOG = pd.DataFrame((df_LOG.loc[(df_LOG['Depth'] <= dmax) & (df_LOG['Depth'] >= dmin)]).mean())
            # new_row = pd.DataFrame(data_imer.loc[(data_imer['Depth'] == list_depth_adcp)]) #extraire la valeur sans aucun traitement : pas de moyennage
            LOG_extract = pd.concat([LOG_extract, new_row_LOG.T], ignore_index=True)

    data_recap['Salinity LOG'] = LOG_extract['Salinity'].values
    if depth_register :
        if len(f_LOG) != 0 :
            list_prof_LOG.append(data_recap['Depth'].iloc[data_recap['Salinity LOG'].last_valid_index()])
            list_sta_LOG.append(station)
    ############################   CTD IMER DATA      "#############################
    col_list_imer = ["Depth", "Temp", "Salinity", "Density", "Chl", "Turbidity"]
    # Done : skip si sheet=station n'existe pas
    wb = load_workbook(file_imer, read_only=True)  # open an Excel file and return a workbook
    if station in wb.sheetnames:
        print('sheet exists in IMER file')
        data_imer = pd.read_excel(file_imer, station, skiprows=23, usecols=col_list_imer)
    else:
        print('Data imer does not exist for this station : ', station)
        data_imer = check_sta_autre_xls(station, wb, file_imer, 23, col_list_imer)
        # TODO IL EST POssible que les données ne soient pas dans le fichier station mais station-X. A chercher
        # il faut créer une panoplie de nom de stations à tester.
        # list_test_sta = [station, station + '-1', station + '-2', station + '-3', station + '-4']
        # data_imer_test = pd.DataFrame(np.nan, index=range(2), columns=data_imer.columns)
        # # j'initialise pour trouver le fichier le plus long dans toutes les possibilités -1 -2 ect
        # for sta in list_test_sta:
        #     if sta in wb.sheetnames:
        #         print(station + ' doesnt exist but ' + sta + ' exists in imer file')
        #         data_imer1 = pd.read_excel(file_imer, sta, skiprows=23, usecols=col_list_imer)
        #         print(data_imer1)
        #         if np.shape(data_imer)[0] > np.shape(data_imer1):
        #             data_imer_test = data_imer1.copy()
        # if np.shape(data_imer_test.shape)[0] < 3:
        #     data_imer = pd.DataFrame(np.nan, index=range(len(list_depth_adcp)), columns=data_imer.columns)
        # else:
        #     data_imer = data_imer_test.copy()
        # data_imer=data_imer.drop(["Turbidity filtered 5"], axis=1)

    turb = signal.medfilt(data_imer['Turbidity'].values, 5)  # calcul de la turbidité avec filter median 5
    data_imer.insert(6, "Turbidity filtered 5", turb)  # insertion de la turb dans le df
    print('data IMER', data_imer)

    # FILLING of the data to check the depth of the different devices
    if depth_register:
        if np.all(np.isnan(data_imer['Salinity'])):  # if there is only NAN values
            print('skip cause all data imer is at Nan')
            # continue
        else:
            if len(list_sta_CTD) == 0:
                list_prof_CTD.append(data_imer['Depth'].max())
                list_sta_CTD.append(station)
            elif len(list_sta_CTD) > 0:
                if list_sta_CTD[-1] == station:  # if previous recorded station is the same, we skip
                    continue
                else:
                    list_prof_CTD.append(data_imer['Depth'].max())
                    list_sta_CTD.append(station)

    ##############     CREATE a file with all values at same depth
    sal_extract = pd.DataFrame()
    D50_extract = pd.DataFrame()
    D90_extract = pd.DataFrame()
    for l in range(len(list_depth_adcp)):
        dmax = list_depth_adcp[l]
        if l != 0:
            dmin = list_depth_adcp[l - 1]
        else:
            dmin = 0
        # print('dmin et max', dmin, dmax) #depth imer : every 0.1m
        new_row = pd.DataFrame((data_imer.loc[(data_imer['Depth'] <= dmax) & (data_imer['Depth'] >= dmin)]).mean())
        # new_row = pd.DataFrame(data_imer.loc[(data_imer['Depth'] == list_depth_adcp)]) #extraire la valeur sans aucun traitement : pas de moyennage
        new_row_D50 = pd.DataFrame(
            (data_unique.loc[(data_unique['Pcorr'] <= dmax) & (data_unique['Pcorr'] >= dmin)]).mean())  # moyennage ?
        D50_extract = pd.concat([D50_extract, new_row_D50.T], ignore_index=True)
        sal_extract = pd.concat([sal_extract, new_row.T], ignore_index=True)
        # print('dmoy', sal_extract['Depth'][l])

    data_recap['Salinity'] = sal_extract['Salinity'].values  # on ajoute les données de salinité
    data_recap['Temp'] = sal_extract['Temp'].values  # on ajoute les données de salinité
    data_recap['Turbidity filtered 5'] = sal_extract['Turbidity filtered 5'].values  # on ajoute les données de salinité
    data_recap["Concentration from turbidity"] = data_recap["Turbidity filtered 5"].values * float(
        dict_adcp[month]['turbTOconcentration'])
    data_recap['Density'] = sal_extract['Density'].values
    data_recap['grad density'] = np.gradient(data_recap['Density'].values,
                                             dict_adcp[month][
                                                 'taille_bin'])  # taillebin=espacement cst entre chq valeur
    data_recap['N2'] = calcul_N2(data_recap['Temp'], data_recap['Salinity']).values
    data_recap['Ri'] = calcul_nb_Ri(data_recap['vitesse north'], data_recap['vitesse east'], data_recap['Density'],
                                    dict_adcp[month]['taille_bin'])
    data_recap["Simpson"] = np.ones(len(data_recap['Depth'])) * float(calcul_simpson(data_recap))
    data_recap['D50'] = D50_extract['D50'].values
    data_recap['D90'] = D50_extract['D90'].values
    data_recap['SPMVC'] = D50_extract['SPMVC'].values
    data_recap['Junge'] = D50_extract['Junge'].values
    # TODO : ajouter nb particule global ? ou de chaque classe ?

    # 23/05/2023 : Calcul de Delta Rhof selon les 2 méthodes
    # Méthode A : si on a les données LISST (SPMVC) filtration ( %OM et SPM) et Density
    liste, liste2 = [], []
    liste.append([data_recap[u].last_valid_index() for u in ['Density', 'D50', '% OM']])
    liste2.append([data_recap[u].last_valid_index() for u in ['Density', 'D50']])

    if any(elem is None for elem in liste[0]):
        # il y a un elem à None, go pour calcul avec Méthode B
        data_recap['Delta rhof methodA'] = np.nan
        data_recap["df"] = 2
    else:
        data_recap['Delta rhof methodA'] = calcul_excess_density_methodA(data_recap)
        print('DELTA RHOF method A ', data_recap['Delta rhof methodA'])
        # Ajouter calcul de df
        data_recap['df'] = calcul_df(data_recap)
    if any(elem is None for elem in liste2[0]):
        # il y a un elem à None, go pour calcul avec Méthode B
        data_recap['Delta rhof methodB'] = np.nan
    else:
        data_recap['Delta rhof methodB'] = calcul_excess_density_methodB(data_recap)
        print('DELTA RHOF method B ', data_recap['Delta rhof methodB'])

    data_recap['ws'] = calcul_ws(data_recap)
    # data_recap['G'] = calcul_G(data_recap)

    print('data_all', data_recap[['Depth', 'vitesse mag', 'Turbidity filtered 5', 'D50']])

    # Je sauvegarde le fichier, à utiliser pour toutes les données surfer.
    if save:
        #path = rep_station + 'Recap_all_param_depth.xlsx'
        path = 'Recap_all_param_' + month + '.xlsx'
        # CONTROLE SI FICHIER EXISTE :
        if not os.path.exists(path):
            # create a new XLSX workbook
            print('I create a new file')
            import xlsxwriter

            wb = xlsxwriter.Workbook(path)  # Create a workbook and add a worksheet.
            if station != station_init:
                worksheet = wb.add_worksheet(station_init)
            else:
                worksheet = wb.add_worksheet(station)
            wb.close()
        else:
            # si la sheet n'existe pas, je la créé
            wb = load_workbook(path, read_only=False)  # open an Excel file and return a workbook
            if station in wb.sheetnames:
                print('sheet1 exists in excel file')
            else:
                if station != station_init:
                    operations_sheet = wb.create_sheet(station_init)
                else:
                    operations_sheet = wb.create_sheet(station)

        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        data_recap.to_excel(writer, sheet_name=station, startrow=0, startcol=0, index=False)  # , header=header)
        writer.save()

        # wb = load_workbook(path, read_only=False)
        # wb.create_sheet(station)
        # data_recap.to_excel(wb, sheet_name=station)
        # wb.save(path)
        # writer = pd.ExcelWriter(path, engine='xlsxwriter')
        # writer.close()
        # read the existing sheets so that openpyxl won't create a new one later

    # # Je sépare en 3 couches : #Done : a faire selon la vraie profondeur ==> On peut dire que c'est ok car on prend la prof max des 3 instruments.
    # if pmax > 4 and not adcp_not_good:  # condition , si non remplie, je ne fais qu'une couche
    #     option = 'adcp'  # 'adcp' ou ' moy' choisi si l'on prend la valeur max des 3 instruement ou si l'on prend la val de adcp
    #     if option == 'moy':
    #         max_depth = max(data_recap['Salinity'].last_valid_index(), data_recap['D50'].last_valid_index(),
    #                         data_recap['vitesse'].last_valid_index())
    #         val_depth = data_recap.loc[max_depth].loc[
    #             'depth']  # c'est la pmax des 3 valeurs, issues de 3 instruments : ADCP, CTD imer, et LISST.
    #     elif option == "adcp":
    #     elif option == "adcp":
    #         val_depth = data_recap.loc[data_recap['vitesse'].last_valid_index()].loc['depth']
    #     nb_couche = 3
    #     list_couche = ['moy_couche' + str(i) for i in range(1, nb_couche + 1)]
    #     for i in range(0, nb_couche):
    #         condition = (
    #                 (data_recap['Depth'] > i * val_depth / 3) & (data_recap['Depth'] < (i + 1) * val_depth / 3))
    #         list_couche[i] = data_recap[condition].copy()
    # elif pmax <= 4:
    #     nb_couche = 1
    #     list_couche = ['moy_couche' + str(i) for i in range(1, nb_couche + 1)]
    #     for i in range(0, nb_couche):
    #         condition = (
    #                 (data_recap['Depth'] > i * val_depth / 3) & (data_recap['Depth'] < (i + 1) * val_depth / 3))
    #         list_couche[i] = data_recap[condition].copy()

print('end')

if depth_register:
    df_depth_ADCP = pd.DataFrame(list_sta_ADCP)
    df_depth_ADCP['max'] = list_prof_ADCP
    df_depth_CTD = pd.DataFrame(list_sta_CTD)
    df_depth_CTD['max'] = list_prof_CTD
    df_depth_LOG = pd.DataFrame(list_sta_LOG)
    df_depth_LOG['max'] = list_prof_LOG
    # df_depth_CTD = df_depth_CTD.drop_duplicates(keep='first')
    # df_depth_ADCP = df_depth_ADCP.drop_duplicates(keep='first')

    rep2 = '/home/penicaud/Documents/Data/'
    file_depth = rep2 + 'Survey_' + month + '/diff_depth_surface_LISST_' + month + '.xlsx'
    df_depth_LISST = pd.read_excel(file_depth)

    df_depth = pd.DataFrame(list_station)
    df_depth = df_depth.set_index(0)

    for sta in df_depth.index:
        print('sta', sta)
        case_ADCP = np.where(sta == df_depth_ADCP[0])[0]
        case_CTD = np.where(sta == df_depth_CTD[0])[0]
        case_LOG = np.where(sta == df_depth_LOG[0])[0]
        case_LISST = np.where(sta == df_depth_LISST['index'])[0]
        # cond = (df_depth['index'] == sta)
        if len(case_ADCP) != 0:  # if sthg corresponds to station i.e len != 0
            df_depth.loc[sta, 'depth ADCP'] = df_depth_ADCP['max'].loc[case_ADCP].values
        if len(case_CTD) != 0:  # if sthg corresponds to station i.e len != 0
            df_depth.loc[sta, 'depth CTD'] = df_depth_CTD['max'].loc[case_CTD].values
        if len(case_LOG) != 0:  # if sthg corresponds to station i.e len != 0
            df_depth.loc[sta, 'depth LOG'] = df_depth_LOG['max'].loc[case_LOG].values
        if len(case_LISST) != 0:  # if sthg corresponds to station i.e len != 0
            df_depth.loc[sta, 'depth LISST'] = df_depth_LISST['max'].loc[case_LISST].values
            for name in ['min down', 'min up', 'height down', 'height up'] :
                df_depth.loc[sta, name] = df_depth_LISST[name].loc[case_LISST].values

    outfile = rep2 + 'Survey_' + month + '/diff_depth_surface_allinstrum_allstations_v2_' + month + '.xlsx'
    df_depth.to_excel(outfile, header=True)

print('end 2')