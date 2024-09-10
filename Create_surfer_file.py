# 26/01/2023 créer un fichier specialement utilisable par surfer
# format : Distance, Depth, Temp, Turb, turb surface (mesuré par turbidimètre),
# turbidity median filter 3, turb median filter 5, transect, station, Date, time, lon lat
#13/04 : je rajoute rho et N2
import csv
import pandas as pd
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
# from sklearn.linear_model import LinearRegression, TheilSenRegressor
# from sklearn.linear_model import RANSACRegressor
from sklearn.metrics import mean_squared_error
import scipy.stats as stats
# import csv
import sys, os, glob
# import xarray as xr
# import matplotlib.colors as mcolors
from matplotlib.ticker import (MultipleLocator, FormatStrFormatter, AutoMinorLocator)
import gsw
import xlsxwriter
import random
import pickle
from itertools import zip_longest
import scipy.signal as signal
from openpyxl import load_workbook

import os
from openpyxl import load_workbook


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    #>>> append_df_to_excel('d:/temp/test.xlsx', df)

    #>>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    #>>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    #>>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

year = '2022'
month = '06'
# day= ['16','17','18']
#day = ['16']  # ,'05'] #ONLY USEFUL FOR 1ST SURVEY 1 5JUNE
suffixe = '.csv'
rep = '/home/penicaud/Documents/Data/CTD/'


#Def variable
index=1
stations_liste=[]
time=[]

if month == '06':
    survey = 'Survey_June'
    file_imer = 'CTD_VU_16-20_06.xlsx'  # Survey to 16-20 june
    #file_imer = 'CTD_VU-I.xlsx'  # survey of 2- JUNE
    # sheets de S1 à S39 puis Fixed S.1 à 25 vérifeir correspondances puis G1 à G8 sauf G7
elif month == '08':
    survey = 'Survey_August'
    file_imer = 'CTD_imer_aug2022.xlsx'
elif month=='10':
    survey='Survey_Octobre'
    file_imer='CTD_imer_octobre.xlsx'

f_imer = rep + survey + '/' + file_imer
print('f_imer', f_imer)

#rep LOG, non utilisé pour le moment -26/01/2023
rep_LOG = rep + survey + '/CTD_LOG/colocalise_CTD/'
liste = glob.glob(rep_LOG + '*' + suffixe)  # Go and find all the files
liste = sorted(liste)
print('len CTD LOG files', len(liste))

#########    JUNE    ########################

if month == '06':
    transect = 'fixed'  # TRANSECT TO CHANGE #T1 T2 T3 T4 or fixed
    # T1 : S1 à S11
    # T2 : S12 à S17
    # T3 : S18 S28
    # T4 : S29 S39
    if transect == 'TJ1':
        sta = range(1, 11)
        i = 0
    elif transect == 'TJ2':
        sta = range(11, 18)
        i = 11
    elif transect == 'TJ3':
        sta = range(18, 29)
        i = 16
    elif transect == 'TJ4':
        sta = range(29, 40)
        i = 27
    elif transect == 'fixed':
        sta = range(1, 26)
        i = 37

    for s in sta:
        if transect == 'fixed':
            stations_liste.append('Fixed S.' + str(s))
        else:
            if s == 12:
                stations_liste.append('S12') #'S12-1'
            elif s == 13:
                print('break 13')
                continue
            elif s == 36:
                print('break 36')
                continue
            else:
                stations_liste.append('S' + str(s))

##############  AUGUST    #######################

elif month == '08':
    transect = 'SF1'  # TRANSECT TO CHANGE #TA1 TA2 TA3 TA4 or SF1 or SF_24
    # TA1 : A1 à A5
    # TA2 : A6 à A12
    # TA3 : A26 A40
    # TA4 : A41 A47
    # SF1 : A13 A25
    # SF_24 : AF1 AF38
    if transect == 'TA1':
        deb = 1
        end = 6  # num station +1 because range(deb,end), end is excluded
        sta = range(deb, end)
        i = deb - 1  # different indice for CTD LOG to scroll the list of files and not beeing affect by the brek 13 36 ect
    elif transect == 'TA2':
        deb = 6
        end = 13
        sta = range(deb, end)
        i = deb - 1
    elif transect == 'TA3':
        deb = 26
        end = 41
        sta = range(deb, end)
        i = deb - 1  # because we only use 25.2
    elif transect == 'TA4':
        deb = 41
        end = 48
        sta = range(deb, end)
        i = deb - 1
    elif transect == 'SF1':
        deb = 13
        end = 26
        sta = range(deb, end)
        i = deb - 1
    elif transect == 'SF_24':
        deb = 1
        end = 39
        sta = range(deb, end)
        i = 47
    else:
        print('ERROR in AUGUST, not good transect or fixed')
        sys.exit(1)

    for s in sta:
        if transect == 'SF_24':
            if s in [5, 8, 10, 14, 21, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35]:  # 14 exists but not on <ctd log
                print('break AF')
                continue
            else:
                stations_liste.append('AF' + str(s))
        else:
            if s == 17:
                print('break 17')
                continue
            if s == 25:
                stations_liste.append('A025-1')
                print('STATION 25')
            else:
                if s < 10:
                    stations_liste.append('A00' + str(s))
                else:
                    stations_liste.append('A0' + str(s))

##############  OCTOBRE    #######################

elif month == '10':
    print('Survey_octobre ')
    transect = 'SFO_24'  # TRANSECT TO CHANGE #TA1 TA2 TA3 TA4 or SF1 or SF_24
    # TO1 : O1 à 07
    # TO2 : 08 à 021
    # TO3 : O22 029
    # TO4 : 030 035
    # SFO_24 : FO1 FO51
    if transect == 'TO1':
        deb = 1
        end = 8  # num station +1 because range(deb,end), end is excluded
        sta = range(deb, end)
        i = deb - 1  # different indice for CTD LOG to scroll the list of files and not beeing affect by the brek 13 36 ect
    elif transect == 'TO2':
        deb = 8
        end = 22
        sta = range(deb, end)
        i = deb - 1
    elif transect == 'TO3':
        deb = 22
        end = 30
        sta = range(deb, end)
        i = deb - 1  # because we only use 25.2
    elif transect == 'TO4':
        deb = 30
        end = 36
        sta = range(deb, end)
        i = deb - 1
    elif transect == 'SFO_24':
        deb = 1
        end = 52
        sta = range(deb, end)
        i = 47 # ATTENTION ??
    else:
        print('ERROR in OCOTBRE, not good transect or fixed')
        sys.exit(1)

    for s in sta:
        #print('boucle s in sta', s, sta)
        if transect == 'SFO_24':
            stations_liste.append('FO' + str(s))
        else:
            if s in [2,4,6]:
                print('break station 2 4 or 6 ')
                continue
            if s == 22:
                stations_liste.append('O22.2')
                print('STATION 22.2')
            else:
                stations_liste.append('O' + str(s))

else:
    print('PB with the file IMER and the transect HERE')
print('Month : ' , survey)
print(' transect and station', transect, stations_liste)

for s in np.arange(0,len(stations_liste)) :
    #for s in sta:
    print('s ',s)
    station=stations_liste[s]
    print('stations_liste[s]', stations_liste[s])
    # open the Imer CTD
    col_list_imer = ["Depth", "Temp", "Salinity", "Density", "Chl", "Turbidity"]
    #read the data of the IMER CTD
    data_imer = pd.read_excel(f_imer, station, skiprows=23,
                              usecols=col_list_imer)  # lambda x : x > 0 and x <= 27 )#, usecols=col_list)
    df_imer = pd.DataFrame(data_imer, columns=["Depth", "Temp", "Salinity", "Density", "Chl", "Turbidity"])
    depth_imer = -df_imer['Depth'] #signe moins pour surfer
    T_imer = df_imer["Temp"]
    S_imer = df_imer["Salinity"]
    D_imer = df_imer["Density"]
    #Chl = df_imer["Chl"]
    Turbidity = df_imer["Turbidity"]
    print('ok pour la lecture de fIMER')

    time_imer=pd.read_excel(f_imer, station, skiprows=13, usecols='A')
    time_imer=time_imer.columns.values[0]
    time_imer=time_imer[10:18]
    print('time', time_imer)
    #time_imer=pd.DataFrame(time_imer, columns=['Time'])
    #sys.exit(1)

    time.append(time_imer)
    print('time' , time)

    # calcule la turbidité avec median filter 3 et 5
    #print('turbidity original', Turbidity)
    Turbidity2 = signal.medfilt(Turbidity) # filter 3
    #print('turb2', Turbidity2)
    Turbidity3 = signal.medfilt(Turbidity, 5) #calcule la turbidité avec median filter 5
    #print('turb3', Turbidity3)
    print('turbidity and filtered turb ok')

    data_Turbidity2=pd.DataFrame(Turbidity2, columns=['Turbidity median filter 3'])
    print(type(data_Turbidity2))
    data_Turbidity3=pd.DataFrame(Turbidity3, columns=['Turbidity median filter 5'])

    print('len turbidity', len(Turbidity)) #sert d'indice pour la ejouter la station suivante à la suite

    #Calcul de N2
    p = 0#1013.25
    SA = gsw.SA_from_SP(S_imer, p, 106.5, 20.5)
    CT = gsw.CT_from_t(SA, T_imer.values, p)
    PT = gsw.pt0_from_t(SA, T_imer.values, p)
    p_n2=np.arange(0,len(Turbidity)*0.001013,0.001013,dtype=float)
    print('len pn2', len(p_n2))
    #tableau qui va de 0 à la surface à la pression de profondeur donnée par len(turbidity) en dm (car avec la ctd imer,
    # une mesure tous les 0.1m, len turbidity donne le nombre de mesures si 54 = 5.4m de prof) *0.001013 dbar car on perd
    #un atm tous les 10m : 0.001013 db tous les 0.1m
    #p_n2=p*p_n2
    [N2, p_mid] = gsw.Nsquared(SA, CT, p_n2, 20.5)
    # ATTENTION §§§ len(p_mid) et len(N2)= len de tous le reste -1 ==> PQ ? Rajouter pour éviter décalage
    N2_bis=np.zeros(len(SA),dtype=float)
    N2_bis[0:len(SA)-1]=N2_bis[0:len(SA)-1]+N2 #BIDOUILLE pour ajoueter un zero à la fin
    N2_bis=pd.DataFrame(N2_bis, columns=['N2 column'])
    print('N2 ok ', len(N2_bis))

    #130423 : J'ajoute les profils de densité
    rho=gsw.rho(SA.values,CT.values,p)
    rho=pd.DataFrame(rho, columns=['rho'])
    #ATTENTION : la CTD mesure aussi la densité.

    ########### JE vais maintenant caler les données de CTD et d'ADCP sur les profondeurs de CTD imer
    # 1. Je vais trouver la station correspondante de imer à ADCP et LISST : Avec l'heure ?
    # 2. aller chercher l'equivalent et la moyenne des X bins dans le fchier ADCP pour le plotter.
    rep_adcp = '/home/penicaud/Documents/Data/ADCP/Survey_' + month + '/'
    rep_LISST = '/home/penicaud/Documents/Data/LISST/Survey_' + month + '/'
    suffixe = '.csv'
    if month == 'June':
        ##############################   DONNEE ADCP #####################################################
        taille_bin = 0.3
        offset = 0.6
        if date.day == 16:
            file_adcp = rep_adcp + str(date.day) + '0' + str(date.month) + '_alldata_BT'
        elif date.day == 17:
            if int(station[1:3]) <= 28:
                file_adcp = rep_adcp + str(date.day) + '0' + str(date.month) + '_T3_alldata_BT'
            else:
                file_adcp = rep_adcp + str(date.day) + '0' + str(date.month) + '_T4_alldata_BT'
        elif date.day == 18:
            file_adcp = rep_adcp + str(date.day) + '0' + str(date.month) + '_alldata_BT'

    #####     ATTENTION A NE PAS ECRASER LES DOCUMENTS
    file_excel='/home/penicaud/Documents/Data/CTD/'+survey+'/'+survey+'_for_surfer_130423.xlsx' #survey_august/TA_for_surfer_copy.xlsx'
    print('file_excel', file_excel, transect)


    #CONTROLE SI FICHIER EXISTE :
    if not os.path.exists(file_excel) :
        # create a new XLSX workbook
        print('I create a new file')
        import xlsxwriter
        wb = xlsxwriter.Workbook(file_excel)         # Create a workbook and add a worksheet.
        worksheet = wb.add_worksheet(transect)
        wb.close()
    else :
        #si la sheet n'existe pas, je la créé
        wb = load_workbook(file_excel, read_only=False)  # open an Excel file and return a workbook
        if transect in wb.sheetnames:
            print('sheet1 exists in excel file')
        else :
            operations_sheet = wb.create_sheet(transect)

    #append_df_to_excel(file_excel, depth_imer, sheet_name=transect, startrow=None,
    #                   truncate_sheet=False, startcol=2 ,header=True)


    # read the existing sheets so that openpyxl won't create a new one later
    book = load_workbook(file_excel)
    writer = pd.ExcelWriter(file_excel, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    if index==1: #ajoute seulement les headers
        header=True
        ind=index-1
    else :
        header=False
        ind=index
        #il faudra ajouter à la main : distance à l'embouchure, turb surface,
    depth_imer.to_excel(writer, sheet_name=transect, startrow=ind, startcol=1, index=False, header=header)
    T_imer.to_excel(writer, sheet_name=transect, startrow=ind, startcol=2, index=False, header=header)
    S_imer.to_excel(writer, sheet_name=transect, startrow=ind, startcol=3, index=False, header=header)
    Turbidity.to_excel(writer, sheet_name=transect, startrow=ind, startcol=4, index=False, header=header)
    data_Turbidity2.to_excel(writer, sheet_name=transect, startrow=ind, startcol=6, index=False, header=header)#, index_label='Turbidity Median Filter 3')
    data_Turbidity3.to_excel(writer, sheet_name=transect, startrow=ind, startcol=7, index=False, header=header)#, index_label='Turbidity Median Filter 3')
    N2_bis.to_excel(writer, sheet_name=transect, startrow=ind, startcol=8, index=False, header=header)
    D_imer.to_excel(writer, sheet_name=transect, startrow=ind, startcol=9, index=False, header=header)
    #TODO : ajouter simpson param
    #TODO : ajouter time, lon et lat, et distance
    index = index + len(Turbidity)
    print('index', index)
    writer.save()
    print('file', file_excel)


    # else :
    #     depth_imer.to_excel(writer, sheet_name=transect, startrow=index, startcol=1, index=False, header=False)
    #     T_imer.to_excel(writer, sheet_name=transect, startrow=index, startcol=2, index=False, header=False)
    #     S_imer.to_excel(writer, sheet_name=transect, startrow=index, startcol=3, index=False, header=False)
    #     Turbidity.to_excel(writer, sheet_name=transect, startrow=index, startcol=4, index=False, header=False)
    #     data_Turbidity2.to_excel(writer, sheet_name=transect, startrow=index, startcol=6, index=False, header=False)#, index_label='Turbidity Median Filter 3')
    #     data_Turbidity3.to_excel(writer, sheet_name=transect, startrow=index, startcol=7, index=False, header=False)#, index_label='Turbidity Median Filter 3')
    #     N2_bis.to_excel(writer, sheet_name=transect, startrow=index, startcol=8, index=False, header=False)
    #     rho.to_excel(writer, sheet_name=transect, startrow=,ind startcol=9, index=False, header=False)

        # To write text
        #wb = xlsxwriter.Workbook(file_excel)
        #existingWorksheet = wb.get_worksheet_by_name(transect)
        #existingWorksheet.write_row(10, index, time_imer)
        #transect.cell(column=11, row=index, value=time_imer)



sys.exit(1)